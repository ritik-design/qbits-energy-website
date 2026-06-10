#!/usr/bin/env python3
"""Extend the Qbits master workbook with v2 additions: 10 new page tabs + 3 ops tabs."""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = "/home/ritik/qbits-energy-website/Marketing/Qbits_Website_Master_Workbook.xlsx"
DST_DESKTOP = "/home/ritik/Desktop/Qbits_Website_Master_Workbook.xlsx"

wb = load_workbook(SRC)

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
P0 = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
P1 = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
P2 = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(top=THIN, left=THIN, right=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")

def add_sheet(name, headers, rows, col_widths=None):
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = BORDER
    for r, row in enumerate(rows, 2):
        for c, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.alignment = WRAP
            cell.border = BORDER
            if c <= len(headers) and headers[c-1] == "Priority":
                if v == "P0": cell.fill = P0
                elif v == "P1": cell.fill = P1
                elif v == "P2": cell.fill = P2
    if col_widths:
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 36

H_STD = ["URL","Meta Title","Target Keyword","Vol (mo)","KD","Intent","Priority","Status","Content Type","Word Count","Schema","Internal Links","Lead Value (INR)","Owner","Notes"]
COLS_STD = [42, 55, 32, 10, 10, 8, 9, 12, 14, 10, 24, 30, 14, 12, 40]

# ========== 19 SOLAR SYSTEM BUNDLES ==========
# Buyers search "X kW solar system price" 3-5x more than "X kW inverter price"
sys_rows = []
sys_data = [
    (1, 60000, 65000, "P0"),
    (2, 110000, 130000, "P0"),
    (3, 160000, 195000, "P0"),
    (5, 250000, 320000, "P0"),
    (6, 290000, 360000, "P1"),
    (8, 380000, 470000, "P1"),
    (10, 470000, 580000, "P0"),
    (15, 700000, 850000, "P1"),
    (20, 920000, 1100000, "P1"),
    (25, 1150000, 1380000, "P2"),
    (50, 2200000, 2750000, "P2"),
    (100, 4200000, 5400000, "P2"),
]
for kw, _, _, prio in sys_data:
    vol_map = {1:3500, 2:3000, 3:6000, 5:9000, 6:1800, 8:1200, 10:5000, 15:2000, 20:2000, 25:1500, 50:1000, 100:700}
    vol = vol_map.get(kw, 800)
    lead = {1:18000, 2:25000, 3:32000, 5:50000, 6:55000, 8:75000, 10:90000, 15:130000, 20:170000, 25:220000, 50:400000, 100:800000}.get(kw, 50000)
    sys_rows.append([
        f"/solar-systems/{kw}-kw-solar-system",
        f"{kw}kW Solar System Price in India 2026 — Complete Cost",
        f"{kw}kw solar system price",
        vol, "Hard" if vol>3000 else "Medium", "C", prio, "Not Started",
        "System Bundle", 1500, "Product+Offer+FAQ",
        f"{kw}kW inverter + panels + battery + use-case", lead, "Content Lead",
        f"{kw}kW system: panels + inverter + structure + install + subsidy net price"
    ])
# Bundle hubs
for tag, kw, vol, prio in [
    ("/solar-systems/on-grid-system-india", "on grid solar system india", 8000, "P0"),
    ("/solar-systems/hybrid-solar-system", "hybrid solar system india", 6000, "P0"),
    ("/solar-systems/off-grid-solar-system", "off grid solar system india", 4500, "P0"),
    ("/solar-systems/solar-system-with-battery", "solar system with battery price india", 4000, "P0"),
    ("/solar-systems/solar-system-for-home", "solar system for home india", 12000, "P0"),
]:
    sys_rows.append([tag, f"{kw.title()} 2026", kw, vol, "Hard", "C", prio, "Not Started",
                     "System Hub", 2000, "Product+ItemList+FAQ", "All wattage system pages", 60000, "Content Lead",
                     "Hub for system category"])
add_sheet("19_Solar_System_Bundles", H_STD, sys_rows, COLS_STD)

# ========== 20 FINANCE & EMI ==========
fin = [
    ("/finance/solar-emi", "Solar EMI in India 2026 — Banks, Rates, Process", "solar emi", 1800, "Medium", "T", "P0", "Master EMI hub", 50000),
    ("/finance/sbi-solar-loan", "SBI Solar Loan — Rates, Eligibility, Apply", "sbi solar loan", 1500, "Medium", "T", "P0", "Top bank", 60000),
    ("/finance/hdfc-solar-loan", "HDFC Bank Solar Loan — 2026 Guide", "hdfc solar loan", 600, "Medium", "T", "P1", "Private bank", 60000),
    ("/finance/icici-solar-loan", "ICICI Bank Solar Loan", "icici solar loan", 500, "Medium", "T", "P1", "Private bank", 60000),
    ("/finance/bajaj-finserv-solar-emi", "Bajaj Finserv Solar EMI — Zero Down Payment", "bajaj finserv solar emi", 800, "Medium", "T", "P0", "NBFC popular for D2C", 55000),
    ("/finance/ireda-solar-loan", "IREDA Solar Loan — Best Rates for Solar", "ireda solar loan", 1200, "Medium", "T", "P0", "Govt agency low rates", 70000),
    ("/finance/canara-bank-solar-loan", "Canara Bank Solar Loan India", "canara bank solar loan", 400, "Easy", "T", "P1", "PSU bank", 50000),
    ("/finance/zero-down-payment-solar", "Zero Down Payment Solar in India", "zero down payment solar", 500, "Easy", "T", "P0", "Big conversion lever", 60000),
    ("/finance/solar-loan-without-collateral", "Solar Loan Without Collateral — Options", "solar loan without collateral", 400, "Easy", "T", "P1", "Affordability gate", 55000),
    ("/finance/opex-vs-capex-solar", "Solar OPEX vs CAPEX — Which Saves More?", "opex vs capex solar", 800, "Medium", "C", "P0", "C&I framing", 200000),
    ("/finance/solar-ppa-india", "Solar PPA Model India — Pros, Cons, Best For", "solar ppa india", 1500, "Medium", "C", "P0", "C&I OPEX", 250000),
    ("/finance/resco-model-solar", "RESCO Solar Model India", "resco solar india", 600, "Medium", "C", "P1", "C&I no-capex", 200000),
]
fin_rows = [[f[0], f[1], f[2], f[3], f[4], f[5], f[6], "Not Started", "Finance Hub", 1800, "WebPage+FAQ", "Calculator+EMI+state subsidy", f[8], "Content Lead", f[7]] for f in fin]
add_sheet("20_Finance_EMI", H_STD, fin_rows, COLS_STD)

# ========== 21 SOLAR FOR TENANTS / NICHE BUYERS ==========
niche = [
    ("/solar/for-tenants", "Solar for Rented Homes & Tenants — What's Possible", "solar for rented home", 800, "Easy", "C", "P0", "Underserved metro buyers", 35000),
    ("/solar/for-builders", "Solar for Builders & Developers — Pre-Wired Homes", "solar for builders", 400, "Easy", "C", "P0", "B2B builder partnerships", 500000),
    ("/solar/portable-solar-home", "Portable Solar Setup for Renters — Plug & Play", "portable solar home india", 300, "Easy", "C", "P1", "Lifestyle-niche", 25000),
    ("/solar/community-solar-india", "Community Solar in India — How It Works", "community solar india", 200, "Easy", "I", "P1", "Apartment/society", 60000),
    ("/solar/heritage-home", "Solar for Heritage / Old Construction Homes", "solar heritage home", 100, "Easy", "C", "P2", "Niche but premium", 80000),
    ("/solar/temple-religious-institutions", "Solar for Temples, Mosques, Gurudwaras", "solar for temple", 250, "Easy", "C", "P1", "Institutional vertical", 80000),
    ("/solar/ngo-csr-projects", "Solar for NGOs & CSR Projects", "solar csr project", 200, "Easy", "C", "P1", "CSR-funded installs", 120000),
    ("/solar/farmhouse-india", "Solar for Farmhouse — Off-Grid & Hybrid", "solar farmhouse india", 600, "Medium", "C", "P0", "Premium off-grid", 80000),
    ("/solar/second-home", "Solar for Vacation / Second Home", "solar second home", 200, "Easy", "C", "P2", "Affluent niche", 80000),
    ("/solar/government-buildings", "Solar for Government Offices & PSU", "solar govt building", 400, "Easy", "C", "P1", "B2G", 500000),
]
niche_rows = [[n[0], n[1], n[2], n[3], n[4], n[5], n[6], "Not Started", "Niche Vertical", 1500, "WebPage+FAQ", "Use-case + city + product", n[8], "Content Lead", n[7]] for n in niche]
add_sheet("21_Niche_Tenants", H_STD, niche_rows, COLS_STD)

# ========== 22 VS DIESEL GENERATOR HUB ==========
dg = [
    ("/solar-vs-diesel-generator", "Solar vs Diesel Generator — Cost & Payback 2026", "solar vs diesel generator", 1200, "Medium", "C", "P0", "Master comparison", 200000),
    ("/solar-vs-diesel-generator-factory", "Solar vs DG for Factory — When to Replace?", "solar vs diesel factory", 400, "Easy", "C", "P0", "Industrial framing", 500000),
    ("/solar-vs-diesel-generator-hospital", "Solar vs DG for Hospital — Reliability Comparison", "solar vs diesel hospital", 200, "Easy", "C", "P1", "Critical load", 400000),
    ("/solar-vs-diesel-generator-construction", "Solar vs DG for Construction Site", "solar vs diesel construction", 150, "Easy", "C", "P1", "Mobile/temporary", 300000),
    ("/diesel-generator-replacement-roi", "Replace DG With Solar — 5-Year ROI Math", "diesel replacement solar roi", 200, "Easy", "C", "P0", "Calculator + content", 400000),
    ("/solar-vs-dg-comparison-table", "Solar vs Diesel Generator — Side-by-Side Table", "solar vs dg comparison", 250, "Easy", "C", "P1", "Comparison data", 250000),
]
dg_rows = [[d[0], d[1], d[2], d[3], d[4], d[5], d[6], "Not Started", "Comparison Hub", 1800, "Article+FAQ+Comparison", "C&I verticals + calculator", d[8], "Content Lead", d[7]] for d in dg]
add_sheet("22_Vs_Diesel_Generator", H_STD, dg_rows, COLS_STD)

# ========== 23 NEAR-ME GEO PAGES ==========
# Generate by pincode cluster — 50 high-density pincode prefixes
near_clusters = [
    ("400-mumbai-suburban", "Mumbai Suburban (400xxx pincodes)", "Mumbai", 600),
    ("400-mumbai-south", "Mumbai South (400001-400035)", "Mumbai", 400),
    ("110-delhi-central", "Delhi Central (110001-110030)", "Delhi", 500),
    ("110-delhi-south", "Delhi South (110017-110049)", "Delhi", 450),
    ("110-delhi-east", "Delhi East", "Delhi", 350),
    ("110-delhi-west", "Delhi West", "Delhi", 350),
    ("110-delhi-north", "Delhi North", "Delhi", 300),
    ("201-noida-greater-noida", "Noida & Greater Noida", "UP/NCR", 800),
    ("122-gurgaon-gurugram", "Gurgaon / Gurugram", "Haryana/NCR", 700),
    ("121-faridabad", "Faridabad", "Haryana", 300),
    ("560-bangalore-east", "Bangalore East (560037-560066)", "Karnataka", 700),
    ("560-bangalore-south", "Bangalore South", "Karnataka", 700),
    ("560-bangalore-north", "Bangalore North", "Karnataka", 500),
    ("560-bangalore-west", "Bangalore West", "Karnataka", 500),
    ("600-chennai-central", "Chennai Central", "Tamil Nadu", 500),
    ("600-chennai-omr", "Chennai OMR / IT Corridor", "Tamil Nadu", 600),
    ("600-chennai-anna-nagar", "Chennai Anna Nagar", "Tamil Nadu", 350),
    ("500-hyderabad-hitec-city", "Hyderabad HITEC City", "Telangana", 500),
    ("500-hyderabad-gachibowli", "Hyderabad Gachibowli", "Telangana", 450),
    ("500-hyderabad-secunderabad", "Hyderabad Secunderabad", "Telangana", 350),
    ("411-pune-baner-aundh", "Pune Baner / Aundh", "Maharashtra", 500),
    ("411-pune-hadapsar", "Pune Hadapsar / Magarpatta", "Maharashtra", 400),
    ("411-pune-wakad-hinjewadi", "Pune Wakad / Hinjewadi", "Maharashtra", 500),
    ("411-pune-kothrud", "Pune Kothrud", "Maharashtra", 300),
    ("700-kolkata-salt-lake", "Kolkata Salt Lake / Sector V", "West Bengal", 350),
    ("700-kolkata-rajarhat", "Kolkata Rajarhat / New Town", "West Bengal", 300),
    ("700-kolkata-south", "Kolkata South", "West Bengal", 250),
    ("380-ahmedabad-bodakdev", "Ahmedabad Bodakdev / SG Highway", "Gujarat", 500),
    ("380-ahmedabad-satellite", "Ahmedabad Satellite / Vastrapur", "Gujarat", 400),
    ("380-ahmedabad-bopal", "Ahmedabad Bopal / South Bopal", "Gujarat", 350),
    ("395-surat-vesu", "Surat Vesu / Pal", "Gujarat", 400),
    ("395-surat-adajan", "Surat Adajan", "Gujarat", 300),
    ("360-rajkot", "Rajkot", "Gujarat", 400),
    ("390-vadodara-alkapuri", "Vadodara Alkapuri / Akota", "Gujarat", 300),
    ("302-jaipur-malviya-nagar", "Jaipur Malviya Nagar / Mansarovar", "Rajasthan", 400),
    ("302-jaipur-vaishali", "Jaipur Vaishali Nagar", "Rajasthan", 350),
    ("226-lucknow-gomti-nagar", "Lucknow Gomti Nagar", "UP", 450),
    ("226-lucknow-hazratganj", "Lucknow Hazratganj", "UP", 300),
    ("208-kanpur", "Kanpur", "UP", 350),
    ("221-varanasi", "Varanasi", "UP", 300),
    ("273-gorakhpur", "Gorakhpur", "UP", 250),
    ("452-indore", "Indore", "MP", 500),
    ("462-bhopal", "Bhopal", "MP", 400),
    ("440-nagpur", "Nagpur", "Maharashtra", 350),
    ("641-coimbatore", "Coimbatore", "Tamil Nadu", 400),
    ("641-tirupur", "Tirupur (Textile Belt)", "Tamil Nadu", 250),
    ("682-kochi-ernakulam", "Kochi Ernakulam", "Kerala", 350),
    ("695-trivandrum", "Trivandrum", "Kerala", 300),
    ("530-vishakhapatnam", "Vishakhapatnam", "AP", 300),
    ("520-vijayawada", "Vijayawada", "AP", 250),
    ("160-chandigarh-mohali", "Chandigarh / Mohali", "Punjab", 400),
]
near_rows = [
    [f"/solar-installer-near-me/{slug}",
     f"Solar Installer Near Me — {label} | Qbits",
     f"solar installer near me {region.lower().split('/')[0].strip()}",
     vol, "Medium", "L", "P0" if vol>=400 else "P1", "Not Started",
     "Near-Me Local", 1200, "LocalBusiness+FAQ+GeoCoordinates",
     "Parent city page + state subsidy + 3 products", 45000, "Content Lead",
     f"Pincode cluster {slug.split('-')[0]}xxx, {region}"]
    for slug, label, region, vol in near_clusters
]
add_sheet("23_Near_Me_Geo", H_STD, near_rows, COLS_STD)

# ========== 24 FAQ HUBS ==========
faq_hubs = [
    ("/faq", "Frequently Asked Questions — Qbits Energy", "qbits faq", 200, "Easy", "I", "P0", "Master FAQ hub"),
    ("/faq/products", "Product FAQs — Solar Inverter Questions Answered", "solar inverter faq", 800, "Medium", "I", "P0", "Voice search + AI Overviews"),
    ("/faq/installation", "Solar Installation FAQs — Everything You Asked", "solar installation faq", 1200, "Medium", "I", "P0", "Pre-purchase intent"),
    ("/faq/subsidy", "PM Surya Ghar & Subsidy FAQs", "pm surya ghar faq", 2500, "Medium", "I", "P0", "Massive subsidy FAQ vol"),
    ("/faq/warranty", "Solar Inverter Warranty FAQs", "solar warranty faq", 500, "Easy", "I", "P0", "Trust"),
    ("/faq/net-metering", "Net Metering FAQs — All DISCOM Questions", "net metering faq", 800, "Medium", "I", "P1", "Process-heavy"),
    ("/faq/maintenance", "Solar Inverter Maintenance FAQs", "solar maintenance faq", 600, "Medium", "I", "P1", "Post-purchase"),
    ("/faq/financing", "Solar Loan & EMI FAQs", "solar loan faq", 500, "Easy", "I", "P0", "Finance gating"),
]
faq_rows = [[f[0], f[1], f[2], f[3], f[4], f[5], f[6], "Not Started", "FAQ Hub", 2200, "FAQPage+Article", "Glossary + product + state subsidy", 25000, "Content Lead", f[7]] for f in faq_hubs]
add_sheet("24_FAQ_Hubs", H_STD, faq_rows, COLS_STD)

# ========== 25 AI/LLM CITATION-OPTIMIZED PAGES ==========
# Pages explicitly written for ChatGPT/Perplexity/Gemini citation
ai_pages = [
    ("/ai-resources/india-solar-glossary-master", "India Solar Energy Glossary — Master Reference", "indian solar glossary", 800, "Medium", "I", "P0", "LLM citation magnet — comprehensive"),
    ("/ai-resources/india-solar-statistics-2026", "India Solar Industry Statistics 2026 — Data", "india solar statistics", 1200, "Medium", "I", "P0", "Stat-heavy citation target"),
    ("/ai-resources/india-discom-directory", "India DISCOM Directory — All 40+ DISCOMs", "india discom list", 1500, "Medium", "I", "P0", "Authority data"),
    ("/ai-resources/india-solar-tariff-database", "India State-wise Electricity Tariff 2026", "electricity tariff state india", 2000, "Hard", "I", "P0", "Cited reference"),
    ("/ai-resources/india-solar-policy-timeline", "India Solar Policy Timeline 1992–2026", "india solar policy history", 500, "Medium", "I", "P1", "Historical depth"),
    ("/ai-resources/almm-list-master", "ALMM List Complete — Phase I, II, III", "almm list complete", 2000, "Medium", "I", "P0", "Master ALMM reference"),
    ("/ai-resources/pm-surya-ghar-stats", "PM Surya Ghar — Live Statistics & Progress", "pm surya ghar statistics", 4000, "Hard", "I", "P0", "Updated monthly — stat magnet"),
    ("/ai-resources/inverter-technical-glossary", "Solar Inverter Technical Glossary — Engineers' Reference", "inverter technical glossary", 400, "Easy", "I", "P0", "Engineer/LLM citation"),
    ("/ai-resources/india-solar-faq-master", "1000 Solar Questions Answered — India Master FAQ", "solar faq india", 1500, "Medium", "I", "P0", "AI Overview seed"),
    ("/ai-resources/india-net-metering-database", "India Net Metering Rules — State-by-State", "net metering rules india", 1500, "Medium", "I", "P0", "Definitive reference"),
    ("/ai-resources/solar-component-bom-india", "Solar System BOM — Complete Component List", "solar system components india", 800, "Medium", "I", "P1", "Buyer reference"),
    ("/ai-resources/iec-standards-solar-inverter", "IEC Standards for Solar Inverters — Indian Context", "iec standards solar", 400, "Easy", "I", "P1", "Spec reference"),
    ("/ai-resources/solar-roi-benchmarks-india", "India Solar ROI Benchmarks by Industry", "solar roi india benchmark", 600, "Medium", "I", "P0", "Data citation"),
    ("/ai-resources/installer-economics-india", "Solar Installer Unit Economics — India", "solar installer economics", 200, "Easy", "I", "P1", "Industry data"),
    ("/ai-resources/india-rooftop-vs-utility", "Rooftop vs Utility-Scale Solar India", "rooftop vs utility solar india", 500, "Medium", "I", "P1", "Sector comparison"),
]
ai_rows = [[a[0], a[1], a[2], a[3], a[4], a[5], a[6], "Not Started", "AI-Citation Page", 3500, "Article+Dataset+FAQ+DefinedTerm",
            "Multiple internal + outbound to MNRE/IEC", 0, "Senior Engineer + Editor",
            f"{a[7]} — write with numbered claims, cited sources, exact stats"] for a in ai_pages]
add_sheet("25_AI_LLM_Citation", H_STD, ai_rows, COLS_STD)

# ========== 26 LEAD MAGNETS (PDF Downloads) ==========
lm = [
    ("/download/solar-buying-guide-2026", "Free India Solar Buying Guide 2026 (PDF)", "solar buying guide pdf", 800, "Medium", "T", "P0", "Master lead magnet"),
    ("/download/sample-solar-proposal", "Sample Solar Proposal — What a Real Quote Looks Like", "sample solar proposal", 500, "Easy", "T", "P0", "Conversion lever — transparency"),
    ("/download/pm-surya-ghar-application-kit", "PM Surya Ghar Application Kit (PDF)", "pm surya ghar application form", 4000, "Hard", "T", "P0", "Massive demand"),
    ("/download/solar-roi-calculator-spreadsheet", "Excel ROI Calculator for Solar — Free Download", "solar roi excel calculator", 400, "Easy", "T", "P0", "B2B-leaning"),
    ("/download/industrial-solar-feasibility-template", "Industrial Solar Feasibility Report Template", "industrial solar template", 200, "Easy", "T", "P1", "C&I lead"),
    ("/download/installer-commissioning-checklist", "Installer Commissioning Checklist (PDF)", "solar commissioning checklist pdf", 300, "Easy", "T", "P1", "B2B installer"),
    ("/download/india-rooftop-solar-index-report", "India Rooftop Solar Index 2026 — Free Report", "india rooftop solar report", 400, "Easy", "T", "P0", "PR-grade lead magnet"),
    ("/download/state-subsidy-comparison-pdf", "State-wise Solar Subsidy Comparison 2026 (PDF)", "state solar subsidy comparison pdf", 300, "Easy", "T", "P1", "Sub vs sub research"),
]
lm_rows = [[l[0], l[1], l[2], l[3], l[4], l[5], l[6], "Not Started", "Lead Magnet", 600, "WebPage+Offer",
            "Gated by email/phone form", 18000, "Content + Design", l[7]] for l in lm]
add_sheet("26_Lead_Magnets", H_STD, lm_rows, COLS_STD)

# ========== 27 BATTERY & STORAGE PARTNERS ==========
bat = [
    ("/batteries/lithium-ion-compatibility", "Lithium-Ion Battery Compatibility — Qbits Hybrid Inverters", "lithium battery solar inverter compatibility", 800, "Medium", "C", "P0", "Hybrid buyer Q"),
    ("/batteries/pylontech-compatibility", "Pylontech Battery + Qbits Hybrid Inverter", "pylontech battery india", 1500, "Medium", "C", "P0", "Top imported battery brand"),
    ("/batteries/exide-solar-battery", "Exide Solar Battery + Qbits Compatibility", "exide solar battery", 2500, "Hard", "C", "P0", "Top Indian brand"),
    ("/batteries/luminous-li-on", "Luminous Li-On Battery + Qbits Hybrid", "luminous li on battery", 800, "Medium", "C", "P1", "Indian Li-ion"),
    ("/batteries/byd-battery-box", "BYD Battery Box + Qbits Hybrid Inverter", "byd battery india", 600, "Medium", "C", "P1", "Premium imported"),
    ("/batteries/lead-acid-vs-lithium", "Lead Acid vs Lithium Solar Battery — 2026 Guide", "lead acid vs lithium solar", 1500, "Medium", "C", "P0", "Decision blog"),
    ("/batteries/solar-battery-sizing", "How to Size a Solar Battery — Step-by-Step", "solar battery sizing guide", 600, "Medium", "C", "P0", "Sizing tool combo"),
    ("/batteries/solar-battery-life-warranty", "Solar Battery Life & Warranty — Honest Look", "solar battery life india", 500, "Medium", "I", "P1", "Lifecycle education"),
]
bat_rows = [[b[0], b[1], b[2], b[3], b[4], b[5], b[6], "Not Started", "Battery", 1800, "Product+FAQ+Comparison",
             "Hybrid product pages + sizing calculator", 65000, "Content Lead", b[7]] for b in bat]
add_sheet("27_Battery_Storage", H_STD, bat_rows, COLS_STD)

# ========== 28 SOLAR 101 BEGINNER FUNNEL ==========
s101 = [
    ("/solar-101", "Solar 101 — Complete Beginner's Guide to Going Solar", "solar 101 india", 800, "Medium", "I", "P0", "Beginner funnel entry"),
    ("/solar-101/what-is-solar-energy", "What Is Solar Energy? — Absolute Beginner's Guide", "what is solar energy", 4000, "Hard", "I", "P0", "Top funnel"),
    ("/solar-101/how-solar-panels-work", "How Do Solar Panels Work? — With Diagrams", "how do solar panels work", 8000, "Hard", "I", "P0", "Foundational"),
    ("/solar-101/types-of-solar-systems", "Types of Solar Systems — On-Grid vs Off-Grid vs Hybrid", "types of solar systems", 3500, "Hard", "I", "P0", "Decision page"),
    ("/solar-101/solar-system-components", "What Components Make Up a Solar System?", "solar system components", 1500, "Medium", "I", "P0", "Reference"),
    ("/solar-101/is-solar-worth-it-india", "Is Solar Worth It in India? — Honest 2026 Analysis", "is solar worth it india", 1500, "Medium", "C", "P0", "Decision-stage"),
    ("/solar-101/myths-about-solar-india", "10 Myths About Solar in India — Debunked", "solar myths india", 800, "Medium", "I", "P0", "Myth-busting"),
    ("/solar-101/your-first-30-days-of-solar", "Your First 30 Days With Solar — What to Expect", "after solar installation", 300, "Easy", "I", "P1", "Post-purchase"),
]
s101_rows = [[s[0], s[1], s[2], s[3], s[4], s[5], s[6], "Not Started", "Beginner Funnel", 2000, "Article+FAQ",
              "Buying guide + subsidy + product hubs", 35000, "Content Lead", s[7]] for s in s101]
add_sheet("28_Solar_101_Beginner", H_STD, s101_rows, COLS_STD)

# ========== OPS 1: CONTENT CALENDAR (52-WEEK) ==========
weeks = []
def w(week, label, p0=None, p1=None, notes=""):
    return [week, label, p0 or "", p1 or "", notes]
# Q1 — Foundation
weeks.append(w("Q1 W1", "Foundation Week", "Core 6 pages live", "", "Hosting+CDN+Schema baseline"))
weeks.append(w("Q1 W2", "Foundation Week", "Top 5 product pages: Q3S, Q5S, QH5, Q10T, QH10T", "", "Schema + WhatsApp CTA"))
weeks.append(w("Q1 W3", "Foundation Week", "Remaining 25 priority products", "", "P0 wattages 1–10kW"))
weeks.append(w("Q1 W4", "Foundation Week", "Subsidy hub: /subsidy/pm-surya-ghar + 10 priority states", "ALMM list page", "Massive traffic driver"))
weeks.append(w("Q1 W5", "Foundation Week", "Tier-1 city pages (8) + Gujarat/MH/TN/UP/KA states", "", "Local SEO core"))
weeks.append(w("Q1 W6", "Foundation Week", "All 12 calculators live", "", "Lead magnets active"))
weeks.append(w("Q1 W7", "BOFU Push", "Best 5kW + Best for Home + Best Hybrid lists", "Best 3kW, 10kW", "BOFU pillars"))
weeks.append(w("Q1 W8", "BOFU Push", "Top 10 Qbits-vs comparison pages", "", "Conversion content"))
weeks.append(w("Q1 W9", "BOFU Push", "Top 10 competitor review pages", "", "E-E-A-T + traffic"))
weeks.append(w("Q1 W10", "Glossary", "30 priority glossary terms (MPPT, ALMM, MNRE, etc.)", "", "Topical authority"))
weeks.append(w("Q1 W11", "Beginner Funnel", "Solar 101 + 6 sub-pages", "", "Top-funnel feeder"))
weeks.append(w("Q1 W12", "AI/USP", "AI/WhatsApp pages + Made-for-India hub (8 pages)", "Solar-systems master 5 pages", "Qbits USP narrative"))
# Q2 — Expansion
weeks.append(w("Q2 W13", "City Sprint", "40 Tier-2 city pages", "", "Local expansion"))
weeks.append(w("Q2 W14", "City Sprint", "Remaining Tier-2 + first 20 Tier-3", "", ""))
weeks.append(w("Q2 W15", "Glossary", "Next 60 glossary terms", "", ""))
weeks.append(w("Q2 W16", "Hindi", "Top 15 Hindi pages launch", "", "Hindi moat begins"))
weeks.append(w("Q2 W17", "Hindi", "Next 15 Hindi pages", "", ""))
weeks.append(w("Q2 W18", "Gujarati", "Top 20 Gujarati pages launch", "", "Home-state moat"))
weeks.append(w("Q2 W19", "Comparisons", "Remaining 30 comparison pages", "", ""))
weeks.append(w("Q2 W20", "Reviews", "Next 25 competitor review pages", "", ""))
weeks.append(w("Q2 W21", "Lists", "Remaining best-lists (20 pages)", "", ""))
weeks.append(w("Q2 W22", "Subsidy", "Remaining 18 state subsidy hubs", "", ""))
weeks.append(w("Q2 W23", "Calculators", "ROI calculator open-source release + outreach", "", "Backlink campaign"))
weeks.append(w("Q2 W24", "Installer Hub", "Installer Academy core 20 pages", "", "B2B moat"))
weeks.append(w("Q2 W25", "Use-Cases", "Top 15 use-case vertical pages", "", "C&I lead gen"))
weeks.append(w("Q2 W26", "Mid-year audit", "Audit, refresh, reprioritize", "", "Half-year checkpoint"))
# Q3 — Scale
weeks.append(w("Q3 W27", "Near-Me", "First 25 near-me pincode pages", "", "Local-SEO scale"))
weeks.append(w("Q3 W28", "Near-Me", "Remaining 25 near-me pages", "", ""))
weeks.append(w("Q3 W29", "Error Codes", "First 40 error-code pages", "", "Support + LLM SEO"))
weeks.append(w("Q3 W30", "Error Codes", "Remaining 40 error pages", "", ""))
weeks.append(w("Q3 W31", "System Bundles", "All 17 system bundle pages", "", "High-volume keywords"))
weeks.append(w("Q3 W32", "Finance", "All 12 finance/EMI pages", "", "Conversion lever"))
weeks.append(w("Q3 W33", "FAQ Hubs", "All 8 FAQ hub pages", "", "Voice + AI search"))
weeks.append(w("Q3 W34", "Lead Magnets", "All 8 PDF lead magnets + gating", "", "Email capture"))
weeks.append(w("Q3 W35", "AI/LLM", "First 8 AI-citation pages", "", "Future-proof SEO"))
weeks.append(w("Q3 W36", "AI/LLM", "Remaining 7 AI-citation pages", "", ""))
weeks.append(w("Q3 W37", "Batteries", "All 8 battery/storage pages", "", "Hybrid story"))
weeks.append(w("Q3 W38", "Verticals", "Remaining 15 use-case verticals", "", ""))
weeks.append(w("Q3 W39", "Niche", "All 10 niche/tenant pages", "", ""))
weeks.append(w("Q3 W40", "vs Diesel", "All 6 vs-DG pages", "", "C&I conversion"))
# Q4 — Editorial scale
weeks.append(w("Q4 W41", "Blog Sprint", "20 case studies", "", "Social proof"))
weeks.append(w("Q4 W42", "Blog Sprint", "20 how-to / buying guide blogs", "", "Top-funnel"))
weeks.append(w("Q4 W43", "Blog Sprint", "20 founder POV / thought leadership", "", "Authority"))
weeks.append(w("Q4 W44", "Hindi Scale", "Next 20 Hindi pages", "", ""))
weeks.append(w("Q4 W45", "Gujarati Scale", "Remaining Gujarati pages", "", ""))
weeks.append(w("Q4 W46", "Tier-3 City", "Next 30 Tier-3 city pages", "", ""))
weeks.append(w("Q4 W47", "Tier-3 City", "Final 50 Tier-3 city pages", "", ""))
weeks.append(w("Q4 W48", "Refresh", "Refresh all Q1 commercial pages", "", "90-day refresh cadence"))
weeks.append(w("Q4 W49", "Refresh", "Refresh subsidy + comparison pages", "", "Annual update"))
weeks.append(w("Q4 W50", "Trust", "All 12 trust/compliance pages", "", "E-E-A-T finalization"))
weeks.append(w("Q4 W51", "Year-end audit", "Performance review + 2027 plan", "", ""))
weeks.append(w("Q4 W52", "Annual report", "India Rooftop Solar Index publish", "", "Year-end link bait"))
add_sheet("OPS_Content_Calendar", ["Week", "Theme", "P0 Deliverable", "P1 Deliverable", "Notes"], weeks, [12, 22, 50, 40, 35])

# ========== OPS 2: KPI TRACKER ==========
kpi = [
    ("Total indexed pages", "—", "200", "500", "850", "1000", "Track via Search Console"),
    ("Referring domains (RDs)", "5–15", "30", "60", "120", "200", "Ahrefs/Semrush"),
    ("Monthly organic clicks", "<2,000", "20,000", "60,000", "150,000", "400,000", "Search Console"),
    ("Branded search volume", "200/mo", "800", "2,500", "5,000", "10,000", "GSC + Google Trends"),
    ("WhatsApp leads / mo", "—", "200", "800", "2,500", "6,000", "WhatsApp Business API"),
    ("Web form leads / mo", "—", "100", "400", "1,200", "3,000", "CRM"),
    ("Calls / mo", "—", "150", "500", "1,500", "4,000", "Call tracking software"),
    ("Quote requests / mo", "—", "100", "400", "1,200", "3,000", "Multi-channel total"),
    ("Online demos / month", "—", "20", "80", "200", "500", "Calendly etc."),
    ("Avg lead-to-quote (days)", "—", "5", "4", "3", "2", "CRM workflow"),
    ("Avg quote-to-order (%)", "—", "8%", "12%", "16%", "20%", "Sales pipeline"),
    ("Orders attributable to organic", "—", "20/mo", "80/mo", "200/mo", "500/mo", "Attribution model"),
    ("Avg order value (residential)", "₹2L", "₹2L", "₹2.1L", "₹2.3L", "₹2.5L", "Sales data"),
    ("Avg order value (C&I)", "₹15L", "₹15L", "₹18L", "₹22L", "₹28L", "Sales data"),
    ("Distributor partner sign-ups", "—", "20", "50", "120", "250", "Channel CRM"),
    ("Installer certifications issued", "—", "30", "100", "300", "750", "LMS"),
    ("Pages rank top-3 (priority KWs)", "—", "10", "40", "150", "400", "Rank tracker"),
    ("Pages rank top-10 (priority KWs)", "—", "30", "120", "400", "800", "Rank tracker"),
    ("Avg LCP (mobile)", "—", "<2.5s", "<2.2s", "<2.0s", "<1.8s", "Core Web Vitals"),
    ("Avg INP (mobile)", "—", "<200ms", "<180ms", "<160ms", "<140ms", "Core Web Vitals"),
    ("AI/LLM citations (ChatGPT/Perplexity)", "—", "10", "50", "200", "500", "Manual sampling"),
    ("YouTube subscribers", "—", "500", "5,000", "20,000", "60,000", "YT Studio"),
    ("Email subscribers", "—", "1,000", "5,000", "20,000", "60,000", "ESP"),
    ("WhatsApp Business subscribers", "—", "2,000", "10,000", "40,000", "120,000", "WhatsApp"),
]
add_sheet("OPS_KPI_Tracker", ["Metric", "Baseline", "Q1 Target", "Q2 Target", "Q3 Target", "Q4 Target", "Measured Via"],
          kpi, [42, 14, 14, 14, 14, 14, 26])

# ========== OPS 3: EDITORIAL OPS (anchor + owner template) ==========
ops3 = [
    ["Internal Anchor Text Bank", "", "", ""],
    ["URL Type", "Preferred Anchor Text Examples", "Never Use", "Notes"],
    ["Product page", "Qbits {model} {kw}kW; the {kw}kW Qbits inverter; this hybrid model", "click here; product link", "Vary, never exact-repeat"],
    ["Best-list page", "best solar inverter list; our top 5kW picks; honest 2026 ranking", "best solar inverter (exact)", "Mix descriptive variations"],
    ["Glossary", "what {term} means; how {term} works; learn more about {term}", "click here", "Soft link"],
    ["Comparison page", "Qbits vs {brand} comparison; side-by-side spec table", "{brand} (alone)", "Anchor cues the comparison"],
    ["Subsidy hub", "PM Surya Ghar guide; subsidy in {state}; how to apply", "subsidy", "Add context"],
    ["City page", "Qbits installers in {city}; solar in {city}", "{city}", "Add intent word"],
    ["Calculator", "use our payback calculator; see your savings", "calculator", "Action word"],
    ["", "", "", ""],
    ["Owner Roles", "", "", ""],
    ["Role", "Responsibility", "Page Types Owned", "Tools"],
    ["Content Lead", "Editorial calendar, briefs, editing", "Blog, glossary, lead magnets", "Notion/Asana"],
    ["Senior Engineer", "Technical accuracy, reviewer credit", "Product, comparison, AI-citation, error code", "Internal review"],
    ["SEO Specialist", "Keyword research, meta, schema, internal links", "All", "Ahrefs, GSC, Rank Math"],
    ["Designer", "Hero images, diagrams, infographics", "All", "Figma"],
    ["Vernacular Editor (Hindi)", "Native Hindi content, no MT", "/hi/* pages", "Direct authoring"],
    ["Vernacular Editor (Gujarati)", "Native Gujarati content", "/gu/* pages", "Direct authoring"],
    ["Outreach Manager", "Link building, PR, guest posts", "Off-page (Link-Building doc)", "Email outreach tool"],
    ["Web Developer", "Schema deploy, CWV optimization", "Technical SEO checklist", "WP, Cloudflare"],
    ["", "", "", ""],
    ["Refresh Cadence", "", "", ""],
    ["Page Type", "Refresh Frequency", "Refresh Triggers", "Owner"],
    ["Subsidy/policy pages", "Quarterly", "Budget, scheme change, court order", "Content Lead"],
    ["Best/list pages", "Every 6 months", "New SKU launch, brand changes", "SEO Specialist"],
    ["Product pages", "On SKU spec change", "Firmware/spec update", "Senior Engineer"],
    ["Comparison pages", "Every 6 months", "Competitor launch", "SEO Specialist"],
    ["Glossary", "Only on standard change", "IEC/BIS update", "Senior Engineer"],
    ["Location pages", "Annually", "Tariff updates, DISCOM change", "Content Lead"],
    ["Blog/case studies", "Annually", "Outdated data", "Content Lead"],
    ["AI-citation pages", "Quarterly", "New official data", "Senior Engineer"],
]
add_sheet("OPS_Editorial_Ops", ["Column 1", "Column 2", "Column 3", "Column 4"], ops3, [32, 60, 28, 30])

# ========== UPDATE README ==========
if "00_README" in wb.sheetnames:
    ws = wb["00_README"]
    # Find last row
    last = ws.max_row + 1
    additions = [
        ("", "", ""),
        ("--- V2 ADDITIONS (added Jun 2026) ---", "", ""),
        ("19_Solar_System_Bundles", "17 system-bundle pages (panel+inverter+battery combos)", "Captures 'XkW solar system price' searches"),
        ("20_Finance_EMI", "12 finance/EMI partner pages", "Conversion lever — affordability gating"),
        ("21_Niche_Tenants", "10 underserved-buyer pages", "Zero-competition niches"),
        ("22_Vs_Diesel_Generator", "6 DG-replacement pages", "C&I conversion narrative"),
        ("23_Near_Me_Geo", "50 pincode-cluster local pages", "Local-SEO domination"),
        ("24_FAQ_Hubs", "8 dedicated FAQ landing pages", "Voice + AI Overview SEO"),
        ("25_AI_LLM_Citation", "15 LLM-citation-optimized pages", "Future of search"),
        ("26_Lead_Magnets", "8 gated PDF downloads", "Email/phone capture"),
        ("27_Battery_Storage", "8 battery partner pages", "Hybrid product story"),
        ("28_Solar_101_Beginner", "8 absolute-beginner funnel pages", "Top-funnel feeder"),
        ("OPS_Content_Calendar", "52-week publishing calendar", "Sprint execution"),
        ("OPS_KPI_Tracker", "Quarterly target tracker", "Measurement discipline"),
        ("OPS_Editorial_Ops", "Anchor text bank + owner roles + refresh cadence", "Editorial operations"),
        ("", "", ""),
        ("TOTAL PAGES (V2)", "1,130 pages mapped", "1,000 original + 130 V2 additions"),
    ]
    for i, row in enumerate(additions):
        for c, v in enumerate(row, 1):
            ws.cell(row=last+i, column=c, value=v).alignment = WRAP

# Save to both locations
wb.save(SRC)
wb.save(DST_DESKTOP)
print(f"SAVED: {SRC}")
print(f"SAVED: {DST_DESKTOP}")
print(f"Total sheets: {len(wb.sheetnames)}")
for s in wb.sheetnames:
    print(f"  - {s}")
