#!/usr/bin/env python3
"""Builds the Qbits Energy master website + keyword workbook with multiple tabs."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
wb.remove(wb.active)

# ---------- Styling helpers ----------
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
PRIORITY_HIGH = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
PRIORITY_MED = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
PRIORITY_LOW = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(top=THIN, left=THIN, right=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")

def add_sheet(name, headers, rows, col_widths=None, freeze="A2"):
    ws = wb.create_sheet(name)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = BORDER
    for r, row in enumerate(rows, 2):
        for c, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.alignment = WRAP
            cell.border = BORDER
            # Priority color
            if c <= len(headers) and headers[c-1] == "Priority":
                if v == "P0": cell.fill = PRIORITY_HIGH
                elif v == "P1": cell.fill = PRIORITY_MED
                elif v == "P2": cell.fill = PRIORITY_LOW
    if col_widths:
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = freeze
    ws.row_dimensions[1].height = 36

# ========== 00 README ==========
readme = [
    ["Tab", "What's in it", "Use it for"],
    ["00_README", "This guide", "Orient yourself"],
    ["01_Core_Pages", "Home, About, Contact, etc.", "MVP launch checklist"],
    ["02_Product_Pages", "60 inverter SKU pages", "Product launch & SEO sprint"],
    ["03_Best_List_Pages", "60 'Best X' commercial pages (BOFU)", "Lead-gen prioritization"],
    ["04_Comparison_Pages", "60 vs-pages (BOFU)", "Conversion content"],
    ["05_Competitor_Reviews", "80 brand/model reviews", "E-E-A-T + comparison link bait"],
    ["06_Glossary_Pages", "150 deep glossary terms", "Topical authority + LLM citations"],
    ["07_City_Pages", "150 city local-SEO pages", "Local lead capture"],
    ["08_State_Pages", "28 state hubs", "State subsidy + local trust"],
    ["09_Blog_Pages", "180 editorial articles", "Top + mid-funnel SEO"],
    ["10_Subsidy_Pages", "32 subsidy/policy hubs", "PM Surya Ghar domination"],
    ["11_Calculator_Tools", "12 interactive tools", "Lead magnets + backlinks"],
    ["12_UseCase_Verticals", "30 vertical pages", "C&I + niche lead capture"],
    ["13_Installer_Hub", "60 B2B installer pages", "Channel loyalty & B2B"],
    ["14_Error_Codes", "80 troubleshooting pages", "Service-call SEO + AI citations"],
    ["15_Hindi_Content", "50 Hindi pages", "Tier-2/3 moat (UP/MP/Bihar/Raj)"],
    ["16_Gujarati_Content", "40 Gujarati pages", "Home-state language moat"],
    ["17_AI_WhatsApp_Special", "10 USP feature pages", "Differentiation moat"],
    ["18_Made_In_India", "15 Atmanirbhar hub pages", "Govt + patriotic-buyer trust"],
    ["KW_BOFU", "Highest-converting bottom-funnel KWs", "Paid + organic lead-gen"],
    ["KW_Commercial", "Mid-funnel commercial KWs", "Comparison content"],
    ["KW_Informational", "Top-funnel education KWs", "Blog editorial calendar"],
    ["KW_Hindi", "Hindi-language KWs", "Hindi content roadmap"],
    ["KW_Subsidy", "Subsidy/policy KWs", "Subsidy hub roadmap"],
    ["KW_Comparison", "Brand-vs-brand KWs", "Comparison page roadmap"],
    ["KW_Geo_Local", "City/state KWs", "Local page roadmap"],
    ["KW_LongTail_Patterns", "Programmatic SEO patterns", "Scale page generation"],
    ["", "", ""],
    ["LEGEND", "", ""],
    ["Priority", "P0=Build in Q1, P1=Q2-Q3, P2=Q4", "Sequence the roadmap"],
    ["Intent", "I=Info, C=Commercial, T=Transactional, N=Nav, L=Local", "Map to funnel"],
    ["KD", "Easy <20 | Medium 20-40 | Hard 40-60 | Very Hard >60", "Allocate effort"],
    ["Vol (mo)", "Indian monthly search volume estimates", "Triangulate; verify with Ahrefs/SEMrush before campaigns"],
    ["Status", "Not Started / In Progress / Published / Updated", "Track execution"],
]
add_sheet("00_README", readme[0], readme[1:], [22, 50, 50])

# ========== 01 CORE PAGES ==========
core = [
    ("/", "Qbits Energy — AI Solar Inverters for India", "solar inverter india", 18000, "Very Hard", "C", "P0", "Not Started", "Homepage", 1200, "Organization+WebSite+BreadcrumbList", "All pillars", "Founder-led hero, USPs, 12-yr warranty, ALMM"),
    ("/about", "About Qbits Energy — Built for India", "qbits energy", 200, "Easy", "I", "P0", "Not Started", "About", 1500, "Organization+AboutPage", "Leadership, factory", "Founder story, mission, manufacturing"),
    ("/contact", "Contact Qbits Energy — Service, Sales, Support", "qbits contact", 100, "Easy", "T", "P0", "Not Started", "Contact", 600, "ContactPage+LocalBusiness", "All locations", "WhatsApp, phone, email, regional offices"),
    ("/dealer-locator", "Find a Qbits Solar Installer Near You", "solar installer near me", 5000, "Hard", "L", "P0", "Not Started", "Tool", 800, "ItemList+LocalBusiness", "All city pages", "Pincode-search map"),
    ("/sitemap", "Sitemap — Qbits Energy", "sitemap", 0, "Easy", "N", "P0", "Not Started", "Utility", 200, "WebPage", "All pages", "HTML sitemap"),
    ("/privacy-policy", "Privacy Policy — Qbits Energy", "privacy policy", 0, "Easy", "I", "P0", "Not Started", "Legal", 1000, "WebPage", "Contact", "GDPR + India DPDP compliant"),
    ("/terms", "Terms of Service — Qbits Energy", "terms", 0, "Easy", "I", "P0", "Not Started", "Legal", 1200, "WebPage", "Privacy", "Standard TOS"),
    ("/warranty", "Qbits 12-Year Warranty Terms", "qbits warranty", 100, "Easy", "I", "P0", "Not Started", "Trust", 1200, "WebPage+FAQPage", "Products", "Plain-English warranty"),
    ("/warranty-registration", "Register Your Qbits Inverter Warranty", "qbits warranty registration", 80, "Easy", "T", "P0", "Not Started", "Tool", 400, "WebPage", "Warranty", "60-second form"),
    ("/certifications", "Qbits Certifications — BIS, ALMM, MNRE, ISO", "almm certified solar inverter", 600, "Medium", "I", "P0", "Not Started", "Trust", 1500, "WebPage", "All product pages", "Cert PDFs displayed"),
    ("/blog", "Qbits Blog — India Solar Insights", "solar blog india", 400, "Medium", "I", "P0", "Not Started", "Hub", 600, "Blog+ItemList", "All blog posts", "Editorial hub"),
    ("/glossary", "Solar Energy Glossary — Qbits Knowledge Base", "solar glossary", 800, "Medium", "I", "P0", "Not Started", "Hub", 800, "DefinedTermSet", "All glossary terms", "Searchable A-Z"),
    ("/calculators", "Free Solar Calculators — Qbits Energy", "solar calculator india", 1500, "Medium", "C", "P0", "Not Started", "Hub", 600, "WebPage", "All calculators", "12 tool index"),
]
headers = ["URL","Meta Title","Target Keyword","Vol (mo)","KD","Intent","Priority","Status","Content Type","Word Count","Schema","Internal Links","Notes"]
add_sheet("01_Core_Pages", headers, core, [38, 55, 30, 10, 10, 8, 9, 14, 14, 11, 32, 30, 50])

# ========== 02 PRODUCT PAGES ==========
products = []
on_grid_single = [(1,"Q1S"),(1.5,"Q1.5S"),(2,"Q2S"),(2.5,"Q2.5S"),(3,"Q3S"),(3.3,"Q3.3S"),(4,"Q4S"),(4.6,"Q4.6S"),(5,"Q5S"),(5.5,"Q5.5S"),(6,"Q6S"),(6.6,"Q6.6S"),(7,"Q7S"),(7.5,"Q7.5S"),(8,"Q8S")]
on_grid_3ph = [(5,"Q5T"),(6,"Q6T"),(8,"Q8T"),(10,"Q10T"),(12,"Q12T"),(15,"Q15T"),(17,"Q17T"),(20,"Q20T"),(25,"Q25T"),(30,"Q30T"),(40,"Q40T"),(50,"Q50T"),(60,"Q60T"),(75,"Q75T"),(100,"Q100T")]
hybrid_single = [(1,"QH1"),(2,"QH2"),(3,"QH3"),(3.6,"QH3.6"),(4,"QH4"),(5,"QH5"),(6,"QH6"),(7,"QH7"),(8,"QH8"),(10,"QH10")]
hybrid_3ph = [(5,"QH5T"),(6,"QH6T"),(8,"QH8T"),(10,"QH10T"),(12,"QH12T"),(15,"QH15T"),(20,"QH20T"),(25,"QH25T"),(30,"QH30T"),(50,"QH50T")]
off_grid = [(1,"QO1"),(1.5,"QO1.5"),(2,"QO2"),(3,"QO3"),(3.5,"QO3.5"),(5,"QO5"),(6,"QO6"),(7.5,"QO7.5"),(10,"QO10"),(12,"QO12")]

def prod_row(kw, model, ptype, slug_type, hot=False):
    url = f"/products/{slug_type}/{str(kw).replace('.','-')}-kw-{model.lower()}"
    kw_str = f"{kw}kw {ptype.lower()} solar inverter"
    vol = {1:3000,2:2500,3:5000,5:8000,10:4000,15:1500,20:1500,25:1000,50:800,100:500}.get(int(kw) if kw==int(kw) else int(kw), 800)
    prio = "P0" if kw in [3,5,10] else ("P1" if kw <= 25 else "P2")
    title = f"Qbits {model} {kw}kW {ptype.title()} Solar Inverter | 12-Yr Warranty"[:60]
    return [url, title, kw_str, vol, "Hard" if vol>3000 else "Medium", "C", prio, "Not Started",
            "Product", 1000, "Product+Review+FAQ", "Category+compare+glossary+city",
            f"{ptype.title()} {kw}kW SKU page; spec table, efficiency curve, install requirements"]

for kw, m in on_grid_single: products.append(prod_row(kw, m, "On-Grid", "on-grid"))
for kw, m in on_grid_3ph: products.append(prod_row(kw, m, "On-Grid 3P", "on-grid-3phase"))
for kw, m in hybrid_single: products.append(prod_row(kw, m, "Hybrid", "hybrid"))
for kw, m in hybrid_3ph: products.append(prod_row(kw, m, "Hybrid 3P", "hybrid-3phase"))
for kw, m in off_grid: products.append(prod_row(kw, m, "Off-Grid", "off-grid"))

add_sheet("02_Product_Pages", headers, products, [40, 55, 30, 10, 10, 8, 9, 14, 12, 11, 24, 30, 50])

# ========== 03 BEST/LIST PAGES (BOFU goldmine) ==========
best_list = [
    # Overall
    ("/best-solar-inverter-india", "Best Solar Inverter in India 2026 — Top 10 Picks", "best solar inverter india", 4500, "Very Hard", "C", "P0", "BOFU pillar; Qbits = Best Value"),
    ("/best-solar-inverter-2026", "Best Solar Inverters of 2026 — Annual Roundup", "best solar inverter 2026", 1500, "Hard", "C", "P0", "Annual refresh"),
    ("/best-solar-inverter-brands-india", "Top Solar Inverter Brands in India (2026)", "best solar inverter brands india", 2500, "Hard", "C", "P0", "Brand ranking"),
    ("/top-10-solar-inverter-india", "Top 10 Solar Inverters in India 2026", "top 10 solar inverter india", 3500, "Hard", "C", "P0", "Listicle SEO"),
    ("/most-reliable-solar-inverter-india", "Most Reliable Solar Inverter Brands India 2026", "most reliable solar inverter india", 800, "Medium", "C", "P1", "Trust-anchored"),
    # By wattage
    ("/best-1kw-solar-inverter-india", "Best 1kW Solar Inverter India 2026", "best 1kw solar inverter india", 600, "Medium", "C", "P1", "Small homes"),
    ("/best-2kw-solar-inverter-india", "Best 2kW Solar Inverter India 2026", "best 2kw solar inverter india", 500, "Medium", "C", "P1", "Compact homes"),
    ("/best-3kw-solar-inverter-india", "Best 3kW Solar Inverter India 2026", "best 3kw solar inverter india", 2000, "Hard", "C", "P0", "Most-common residential"),
    ("/best-5kw-solar-inverter-india", "Best 5kW Solar Inverter India 2026", "best 5kw solar inverter india", 4000, "Hard", "C", "P0", "Highest-volume wattage"),
    ("/best-6kw-solar-inverter-india", "Best 6kW Solar Inverter India 2026", "best 6kw solar inverter india", 1000, "Medium", "C", "P1", "Large home"),
    ("/best-10kw-solar-inverter-india", "Best 10kW Solar Inverter India 2026", "best 10kw solar inverter india", 2500, "Hard", "C", "P0", "Bridge home/SMB"),
    ("/best-15kw-solar-inverter-india", "Best 15kW Solar Inverter India 2026", "best 15kw solar inverter india", 800, "Medium", "C", "P1", "Small commercial"),
    ("/best-25kw-solar-inverter-india", "Best 25kW Solar Inverter India 2026", "best 25kw solar inverter india", 600, "Medium", "C", "P1", "MSME"),
    ("/best-50kw-solar-inverter-india", "Best 50kW Solar Inverter India 2026", "best 50kw solar inverter india", 500, "Medium", "C", "P2", "Mid-commercial"),
    ("/best-100kw-solar-inverter-india", "Best 100kW Solar Inverter India 2026", "best 100kw solar inverter india", 400, "Medium", "C", "P2", "Industrial"),
    # By type
    ("/best-hybrid-solar-inverter-india", "Best Hybrid Solar Inverter India 2026", "best hybrid solar inverter india", 2500, "Hard", "C", "P0", "Power-cut buyers"),
    ("/best-on-grid-solar-inverter-india", "Best On-Grid Solar Inverter India 2026", "best on grid solar inverter india", 2500, "Hard", "C", "P0", "Subsidy buyers"),
    ("/best-off-grid-solar-inverter-india", "Best Off-Grid Solar Inverter India 2026", "best off grid solar inverter india", 1500, "Hard", "C", "P1", "Rural"),
    ("/best-3-phase-solar-inverter-india", "Best 3-Phase Solar Inverter India 2026", "best 3 phase solar inverter india", 800, "Medium", "C", "P1", "Commercial"),
    ("/best-single-phase-solar-inverter-india", "Best Single-Phase Solar Inverter India 2026", "best single phase solar inverter india", 600, "Medium", "C", "P1", "Residential"),
    ("/best-string-inverter-india", "Best String Inverter India 2026", "best string inverter india", 500, "Medium", "C", "P1", "Technical buyers"),
    ("/best-micro-inverter-india", "Best Micro Inverter India 2026", "best micro inverter india", 400, "Medium", "C", "P2", "Niche but high intent"),
    ("/best-transformerless-solar-inverter", "Best Transformerless Solar Inverter India", "best transformerless solar inverter", 300, "Easy", "C", "P2", "Spec-led buyer"),
    # By use case
    ("/best-solar-inverter-for-home", "Best Solar Inverter for Home India 2026", "best solar inverter for home", 6000, "Very Hard", "C", "P0", "Highest-volume BOFU"),
    ("/best-solar-inverter-for-3-bhk", "Best Solar Inverter for 3 BHK Home", "best solar inverter for 3 bhk", 1500, "Medium", "C", "P0", "Demographic-targeted"),
    ("/best-solar-inverter-for-villa", "Best Solar Inverter for Villa & Bungalow", "best solar inverter for villa", 600, "Medium", "C", "P1", "Premium"),
    ("/best-solar-inverter-for-shop", "Best Solar Inverter for Shop & Retail", "best solar inverter for shop", 1000, "Medium", "C", "P1", "SMB"),
    ("/best-solar-inverter-for-office", "Best Solar Inverter for Office Space", "best solar inverter for office", 800, "Medium", "C", "P1", "SMB"),
    ("/best-solar-inverter-for-factory", "Best Solar Inverter for Factory & Industry", "best solar inverter for factory", 1200, "Medium", "C", "P0", "C&I"),
    ("/best-solar-inverter-for-school", "Best Solar Inverter for School", "best solar inverter for school", 500, "Easy", "C", "P1", "Institutional"),
    ("/best-solar-inverter-for-hospital", "Best Solar Inverter for Hospital", "best solar inverter for hospital", 400, "Easy", "C", "P1", "Critical-load"),
    ("/best-solar-inverter-for-petrol-pump", "Best Solar Inverter for Petrol Pump", "best solar inverter for petrol pump", 500, "Easy", "C", "P1", "Vertical niche"),
    ("/best-solar-inverter-for-water-pump", "Best Solar Inverter for Water Pump (KUSUM)", "best solar inverter for water pump", 1500, "Medium", "C", "P0", "KUSUM scheme"),
    ("/best-solar-inverter-for-ev-charging", "Best Solar Inverter for EV Charging Home", "best solar inverter for ev charging", 500, "Easy", "C", "P1", "Future-proof angle"),
    ("/best-on-grid-inverter-for-commercial", "Best On-Grid Inverter for Commercial Use", "best on grid inverter commercial", 600, "Medium", "C", "P1", "B2B"),
    # By price
    ("/best-solar-inverter-under-30000", "Best Solar Inverter Under ₹30,000", "best solar inverter under 30000", 1200, "Medium", "C", "P0", "Budget"),
    ("/best-solar-inverter-under-50000", "Best Solar Inverter Under ₹50,000", "best solar inverter under 50000", 1500, "Medium", "C", "P0", "Budget"),
    ("/best-solar-inverter-under-100000", "Best Solar Inverter Under ₹1 Lakh", "best solar inverter under 100000", 800, "Medium", "C", "P1", "Mid"),
    ("/best-budget-solar-inverter-india", "Best Budget Solar Inverter India 2026", "best budget solar inverter india", 800, "Medium", "C", "P1", "Value seekers"),
    ("/best-premium-solar-inverter-india", "Best Premium Solar Inverter India 2026", "best premium solar inverter india", 400, "Easy", "C", "P2", "Premium seekers"),
    # By feature
    ("/best-ai-solar-inverter-india", "Best AI-Powered Solar Inverter in India", "best ai solar inverter india", 300, "Easy", "C", "P0", "Qbits USP — own it"),
    ("/best-wifi-solar-inverter-india", "Best Wi-Fi Smart Solar Inverter India", "best wifi solar inverter india", 1000, "Medium", "C", "P0", "Smart home"),
    ("/best-lithium-compatible-solar-inverter", "Best Lithium-Compatible Solar Inverter India", "best lithium solar inverter india", 600, "Medium", "C", "P1", "Battery future"),
    ("/best-solar-inverter-with-longest-warranty", "Solar Inverters With the Longest Warranty in India", "best warranty solar inverter india", 400, "Easy", "C", "P0", "Qbits 12-yr USP"),
    ("/best-solar-inverter-with-app-monitoring", "Best App-Controlled Solar Inverter India", "best app solar inverter india", 500, "Medium", "C", "P1", "Tech buyers"),
    # By state
    ("/best-solar-inverter-gujarat", "Best Solar Inverter in Gujarat 2026", "best solar inverter gujarat", 1500, "Medium", "L", "P0", "Home state"),
    ("/best-solar-inverter-maharashtra", "Best Solar Inverter in Maharashtra 2026", "best solar inverter maharashtra", 1500, "Medium", "L", "P0", "Top market"),
    ("/best-solar-inverter-tamil-nadu", "Best Solar Inverter in Tamil Nadu 2026", "best solar inverter tamil nadu", 1200, "Medium", "L", "P0", "C&I heavy"),
    ("/best-solar-inverter-karnataka", "Best Solar Inverter in Karnataka 2026", "best solar inverter karnataka", 1000, "Medium", "L", "P1", "Bangalore tech"),
    ("/best-solar-inverter-uttar-pradesh", "Best Solar Inverter in Uttar Pradesh 2026", "best solar inverter uttar pradesh", 1200, "Medium", "L", "P0", "Growth market"),
    ("/best-solar-inverter-rajasthan", "Best Solar Inverter in Rajasthan 2026", "best solar inverter rajasthan", 800, "Medium", "L", "P1", "High DNI"),
    ("/best-solar-inverter-kerala", "Best Solar Inverter in Kerala 2026", "best solar inverter kerala", 700, "Medium", "L", "P1", "Monsoon angle"),
    ("/best-solar-inverter-delhi", "Best Solar Inverter in Delhi NCR 2026", "best solar inverter delhi", 2500, "Hard", "L", "P0", "Top metro"),
    ("/best-solar-inverter-punjab", "Best Solar Inverter in Punjab 2026", "best solar inverter punjab", 600, "Medium", "L", "P1", "Ag + residential"),
    ("/best-solar-inverter-haryana", "Best Solar Inverter in Haryana 2026", "best solar inverter haryana", 600, "Medium", "L", "P1", "NCR fringe"),
    ("/best-solar-inverter-andhra-pradesh", "Best Solar Inverter in Andhra Pradesh 2026", "best solar inverter andhra pradesh", 600, "Medium", "L", "P1", "Industrial"),
    ("/best-solar-inverter-telangana", "Best Solar Inverter in Telangana 2026", "best solar inverter telangana", 700, "Medium", "L", "P1", "Hyderabad C&I"),
    ("/best-solar-inverter-west-bengal", "Best Solar Inverter in West Bengal 2026", "best solar inverter west bengal", 500, "Medium", "L", "P2", "WBSEDCL"),
    ("/best-solar-inverter-madhya-pradesh", "Best Solar Inverter in Madhya Pradesh 2026", "best solar inverter madhya pradesh", 500, "Medium", "L", "P1", "Growing"),
    ("/best-solar-inverter-odisha", "Best Solar Inverter in Odisha 2026", "best solar inverter odisha", 400, "Easy", "L", "P2", "Emerging"),
]
bl_headers = ["URL","Meta Title","Target Keyword","Vol (mo)","KD","Intent","Priority","Notes"]
add_sheet("03_Best_List_Pages", bl_headers, best_list, [42, 50, 32, 10, 10, 8, 9, 50])

# ========== 04 COMPARISON PAGES ==========
comparisons = [
    # Qbits vs major brands
    ("/compare/qbits-vs-sungrow-solar-inverter", "Qbits vs Sungrow — Honest 2026 Comparison", "qbits vs sungrow", 100, "Easy", "C", "P0", "BOFU vs Tier-1 import"),
    ("/compare/qbits-vs-growatt-solar-inverter", "Qbits vs Growatt — Specs, Price, Verdict", "qbits vs growatt", 150, "Easy", "C", "P0", "Highest-search competitor"),
    ("/compare/qbits-vs-luminous-solar-inverter", "Qbits vs Luminous Solar — Detailed Comparison", "qbits vs luminous", 200, "Easy", "C", "P0", "Indian household competitor"),
    ("/compare/qbits-vs-microtek-solar-inverter", "Qbits vs Microtek — Which Is Better 2026", "qbits vs microtek", 100, "Easy", "C", "P0", "Budget competitor"),
    ("/compare/qbits-vs-havells-solar-inverter", "Qbits vs Havells Solar — 2026 Comparison", "qbits vs havells", 80, "Easy", "C", "P1", "Premium Indian"),
    ("/compare/qbits-vs-solis-solar-inverter", "Qbits vs Solis Inverter — Detailed Comparison", "qbits vs solis", 80, "Easy", "C", "P1", "String specialist"),
    ("/compare/qbits-vs-deye-solar-inverter", "Qbits vs Deye Hybrid Inverter — 2026", "qbits vs deye", 100, "Easy", "C", "P0", "Rising hybrid competitor"),
    ("/compare/qbits-vs-goodwe-solar-inverter", "Qbits vs GoodWe — Comparison & Verdict", "qbits vs goodwe", 60, "Easy", "C", "P1", "Premium hybrid"),
    ("/compare/qbits-vs-utl-solar-inverter", "Qbits vs UTL Solar — 2026 Comparison", "qbits vs utl", 80, "Easy", "C", "P1", "Indian budget-hybrid"),
    ("/compare/qbits-vs-waaree-solar-inverter", "Qbits vs Waaree Solar — Comparison", "qbits vs waaree", 100, "Easy", "C", "P0", "Top Indian brand"),
    ("/compare/qbits-vs-adani-solar-inverter", "Qbits vs Adani Solar — Comparison", "qbits vs adani", 60, "Easy", "C", "P1", "Adani brand draw"),
    ("/compare/qbits-vs-polycab-solar-inverter", "Qbits vs Polycab Solar — Comparison", "qbits vs polycab", 50, "Easy", "C", "P2", "New entrant"),
    ("/compare/qbits-vs-v-guard-solar-inverter", "Qbits vs V-Guard Solar — Comparison", "qbits vs v-guard", 50, "Easy", "C", "P2", "Electrical brand"),
    ("/compare/qbits-vs-su-kam-solar-inverter", "Qbits vs Su-Kam Solar — Comparison", "qbits vs sukam", 60, "Easy", "C", "P2", "Legacy brand"),
    ("/compare/qbits-vs-exide-solar-inverter", "Qbits vs Exide Solar — Comparison", "qbits vs exide", 50, "Easy", "C", "P2", "Battery brand entry"),
    ("/compare/qbits-vs-huawei-solar-inverter", "Qbits vs Huawei FusionSolar — Comparison", "qbits vs huawei", 60, "Easy", "C", "P2", "Premium Chinese"),
    ("/compare/qbits-vs-delta-solar-inverter", "Qbits vs Delta Solar — Comparison", "qbits vs delta", 40, "Easy", "C", "P2", "Taiwan tier-1"),
    ("/compare/qbits-vs-fronius-solar-inverter", "Qbits vs Fronius Solar — Comparison", "qbits vs fronius", 50, "Easy", "C", "P2", "Austrian premium"),
    ("/compare/qbits-vs-sma-solar-inverter", "Qbits vs SMA Solar — Comparison", "qbits vs sma", 50, "Easy", "C", "P2", "German premium"),
    ("/compare/qbits-vs-ksolare-solar-inverter", "Qbits vs K'Solare — Comparison", "qbits vs ksolare", 30, "Easy", "C", "P2", "Indian challenger"),
    # Qbits vs wattage-specific
    ("/compare/qbits-vs-growatt-5kw", "Qbits vs Growatt 5kW — Which One Wins?", "qbits vs growatt 5kw", 80, "Easy", "C", "P0", "Hottest SKU comparison"),
    ("/compare/qbits-vs-growatt-10kw", "Qbits vs Growatt 10kW — Comparison", "qbits vs growatt 10kw", 50, "Easy", "C", "P1", "Bridge segment"),
    ("/compare/qbits-vs-sungrow-5kw", "Qbits vs Sungrow 5kW — Comparison", "qbits vs sungrow 5kw", 60, "Easy", "C", "P0", "Tier-1 vs India"),
    ("/compare/qbits-vs-sungrow-10kw", "Qbits vs Sungrow 10kW — Comparison", "qbits vs sungrow 10kw", 40, "Easy", "C", "P1", "Commercial"),
    ("/compare/qbits-vs-solis-5kw", "Qbits vs Solis 5kW — Comparison", "qbits vs solis 5kw", 40, "Easy", "C", "P1", "String specialist"),
    ("/compare/qbits-vs-deye-5kw", "Qbits vs Deye 5kW Hybrid — Comparison", "qbits vs deye 5kw", 60, "Easy", "C", "P0", "Hot hybrid SKU"),
    ("/compare/qbits-vs-deye-10kw", "Qbits vs Deye 10kW Hybrid — Comparison", "qbits vs deye 10kw", 40, "Easy", "C", "P1", "Hybrid 3-phase"),
    ("/compare/qbits-vs-luminous-3kw", "Qbits vs Luminous 3kW — Comparison", "qbits vs luminous 3kw", 80, "Easy", "C", "P1", "Residential entry"),
    ("/compare/qbits-vs-luminous-5kw", "Qbits vs Luminous 5kW — Comparison", "qbits vs luminous 5kw", 80, "Easy", "C", "P0", "Volume SKU"),
    ("/compare/qbits-vs-microtek-3kw", "Qbits vs Microtek 3kW — Comparison", "qbits vs microtek 3kw", 60, "Easy", "C", "P1", "Budget vs Qbits"),
    ("/compare/qbits-vs-utl-5kw", "Qbits vs UTL 5kW — Comparison", "qbits vs utl 5kw", 50, "Easy", "C", "P1", "Hybrid clash"),
    ("/compare/qbits-vs-havells-5kw", "Qbits vs Havells 5kW — Comparison", "qbits vs havells 5kw", 40, "Easy", "C", "P2", "Premium Indian"),
    ("/compare/qbits-vs-goodwe-5kw", "Qbits vs GoodWe 5kW — Comparison", "qbits vs goodwe 5kw", 40, "Easy", "C", "P2", "Hybrid premium"),
    ("/compare/qbits-vs-waaree-5kw", "Qbits vs Waaree 5kW — Comparison", "qbits vs waaree 5kw", 50, "Easy", "C", "P1", "Indian volume"),
    ("/compare/qbits-vs-adani-5kw", "Qbits vs Adani 5kW — Comparison", "qbits vs adani 5kw", 40, "Easy", "C", "P2", "Adani brand pull"),
    # Brand vs Brand (high-volume non-Qbits)
    ("/compare/sungrow-vs-growatt-solar-inverter", "Sungrow vs Growatt — Honest Comparison 2026", "sungrow vs growatt", 1200, "Medium", "C", "P0", "Authority play — capture both audiences"),
    ("/compare/growatt-vs-solis-solar-inverter", "Growatt vs Solis — Detailed Comparison", "growatt vs solis", 700, "Medium", "C", "P0", "High intent"),
    ("/compare/sungrow-vs-solis-solar-inverter", "Sungrow vs Solis — 2026 Comparison", "sungrow vs solis", 600, "Medium", "C", "P0", "Premium clash"),
    ("/compare/luminous-vs-microtek-solar-inverter", "Luminous vs Microtek Solar — Comparison", "luminous vs microtek solar", 1500, "Medium", "C", "P0", "Indian mass-market"),
    ("/compare/luminous-vs-sukam-solar-inverter", "Luminous vs Su-Kam Solar — Comparison", "luminous vs sukam", 800, "Medium", "C", "P1", "Legacy clash"),
    ("/compare/luminous-vs-havells-solar-inverter", "Luminous vs Havells Solar — Comparison", "luminous vs havells", 700, "Medium", "C", "P1", "Indian premium"),
    ("/compare/microtek-vs-sukam-solar-inverter", "Microtek vs Su-Kam Solar — Comparison", "microtek vs sukam", 500, "Medium", "C", "P1", "Budget clash"),
    ("/compare/havells-vs-polycab-solar-inverter", "Havells vs Polycab Solar — Comparison", "havells vs polycab", 400, "Medium", "C", "P2", "Electrical entrants"),
    ("/compare/deye-vs-growatt-solar-inverter", "Deye vs Growatt Hybrid — Comparison", "deye vs growatt", 600, "Medium", "C", "P0", "Hybrid clash"),
    ("/compare/waaree-vs-adani-solar-inverter", "Waaree vs Adani Solar — Comparison", "waaree vs adani solar", 700, "Medium", "C", "P0", "Indian mfg clash"),
    ("/compare/goodwe-vs-growatt-solar-inverter", "GoodWe vs Growatt — Comparison", "goodwe vs growatt", 500, "Medium", "C", "P1", "Hybrid spec"),
    ("/compare/utl-vs-luminous-solar-inverter", "UTL vs Luminous Solar — Comparison", "utl vs luminous", 600, "Medium", "C", "P1", "Indian hybrid"),
    ("/compare/fronius-vs-sma-solar-inverter", "Fronius vs SMA Solar — Comparison", "fronius vs sma", 400, "Medium", "C", "P2", "Premium imports"),
    ("/compare/huawei-vs-sungrow-solar-inverter", "Huawei vs Sungrow Solar — Comparison", "huawei vs sungrow", 400, "Medium", "C", "P2", "Chinese giants"),
    ("/compare/enphase-vs-solaredge-microinverter", "Enphase vs SolarEdge Microinverter", "enphase vs solaredge", 500, "Medium", "C", "P2", "Microinverter niche"),
    # Type vs Type
    ("/compare/hybrid-vs-on-grid-inverter", "Hybrid vs On-Grid Inverter — Which to Choose?", "hybrid vs on grid inverter", 2200, "Hard", "I/C", "P0", "Foundational decision"),
    ("/compare/on-grid-vs-off-grid-inverter", "On-Grid vs Off-Grid Inverter — Explained", "on grid vs off grid inverter", 3500, "Hard", "I/C", "P0", "Highest-vol type comparison"),
    ("/compare/string-vs-micro-inverter", "String vs Micro Inverter — Which Is Better?", "string vs micro inverter", 1800, "Hard", "I", "P1", "Tech buyer"),
    ("/compare/string-vs-central-inverter", "String vs Central Inverter — Comparison", "string vs central inverter", 1200, "Medium", "I", "P1", "Utility-scale"),
    ("/compare/mppt-vs-pwm-inverter", "MPPT vs PWM Inverter — Complete Guide", "mppt vs pwm", 1500, "Medium", "I", "P0", "Tech foundation"),
    ("/compare/single-phase-vs-3-phase-inverter", "Single-Phase vs 3-Phase Solar Inverter", "single phase vs 3 phase inverter", 800, "Medium", "I", "P1", "Connection type"),
    ("/compare/transformer-vs-transformerless-inverter", "Transformer vs Transformerless Inverter", "transformer vs transformerless", 600, "Medium", "I", "P2", "Spec buyer"),
    ("/compare/grid-tie-vs-off-grid-inverter", "Grid-Tie vs Off-Grid Inverter Explained", "grid tie vs off grid", 1200, "Medium", "I", "P1", "Synonym-cluster"),
    ("/compare/solar-inverter-vs-ups", "Solar Inverter vs UPS — What's the Difference?", "solar inverter vs ups", 1200, "Medium", "I", "P0", "Beginner confusion"),
    ("/compare/5kw-vs-10kw-inverter", "5kW vs 10kW Solar Inverter — Sizing Guide", "5kw vs 10kw inverter", 400, "Easy", "I", "P1", "Sizing question"),
]
cmp_headers = ["URL","Meta Title","Target Keyword","Vol (mo)","KD","Intent","Priority","Notes"]
add_sheet("04_Comparison_Pages", cmp_headers, comparisons, [50, 50, 30, 10, 10, 8, 9, 50])

# ========== 05 COMPETITOR REVIEWS ==========
reviews = [
    # Sungrow 5
    ("/reviews/sungrow-solar-inverter-review","Sungrow Solar Inverter Review India 2026 — Honest","sungrow inverter review",1500,"Medium","C","P0"),
    ("/reviews/sungrow-sg5ktl-mt-review","Sungrow SG5KTL-MT 5kW Review","sungrow sg5ktl mt review",300,"Easy","C","P1"),
    ("/reviews/sungrow-sh5k-review","Sungrow SH5K Hybrid Review","sungrow sh5k review",250,"Easy","C","P1"),
    ("/reviews/sungrow-sh10rt-review","Sungrow SH10RT Hybrid Review","sungrow sh10rt review",200,"Easy","C","P2"),
    ("/reviews/sungrow-sg110cx-review","Sungrow SG110CX Commercial Review","sungrow sg110cx review",150,"Easy","C","P2"),
    # Growatt 6
    ("/reviews/growatt-solar-inverter-review","Growatt Solar Inverter Review India 2026","growatt review",2500,"Medium","C","P0"),
    ("/reviews/growatt-mic-3000tl-review","Growatt MIC 3000TL Review","growatt mic 3000tl review",500,"Easy","C","P0"),
    ("/reviews/growatt-min-5000tl-x-review","Growatt MIN 5000TL-X Review","growatt min 5000tl x review",400,"Easy","C","P1"),
    ("/reviews/growatt-spf-5000-review","Growatt SPF 5000 Off-Grid Review","growatt spf 5000 review",350,"Easy","C","P1"),
    ("/reviews/growatt-sph-6000-review","Growatt SPH 6000 Hybrid Review","growatt sph 6000 review",300,"Easy","C","P1"),
    ("/reviews/growatt-max-50ktl3-lv-review","Growatt MAX 50KTL3-LV Review","growatt max 50ktl3 review",150,"Easy","C","P2"),
    # Deye 4
    ("/reviews/deye-solar-inverter-review","Deye Solar Inverter Review India 2026","deye inverter review",1200,"Medium","C","P0"),
    ("/reviews/deye-sun-5k-sg03lp1-review","Deye SUN-5K-SG03LP1 Review","deye 5k review",300,"Easy","C","P1"),
    ("/reviews/deye-sun-10k-sg04lp3-review","Deye SUN-10K-SG04LP3 Review","deye 10k review",250,"Easy","C","P1"),
    ("/reviews/deye-sun-12k-sg04lp3-review","Deye SUN-12K-SG04LP3 Review","deye 12k review",200,"Easy","C","P2"),
    # Luminous 5
    ("/reviews/luminous-solar-inverter-review","Luminous Solar Inverter Review India 2026","luminous solar review",2500,"Medium","C","P0"),
    ("/reviews/luminous-solarverter-pro-review","Luminous Solarverter Pro Review","luminous solarverter review",400,"Easy","C","P1"),
    ("/reviews/luminous-nxg-1850-review","Luminous NXG 1850 Review","luminous nxg 1850 review",350,"Easy","C","P1"),
    ("/reviews/luminous-shine-1250-review","Luminous Shine 1250 Review","luminous shine 1250 review",300,"Easy","C","P2"),
    ("/reviews/luminous-solar-grid-tie-review","Luminous Solar Grid-Tie Inverter Review","luminous grid tie review",200,"Easy","C","P2"),
    # Microtek 4
    ("/reviews/microtek-solar-inverter-review","Microtek Solar Inverter Review India 2026","microtek solar review",1500,"Medium","C","P0"),
    ("/reviews/microtek-mppt-hybrid-review","Microtek MPPT Hybrid Review","microtek mppt hybrid review",300,"Easy","C","P1"),
    ("/reviews/microtek-super-solar-review","Microtek Super Solar PCU Review","microtek super solar review",250,"Easy","C","P2"),
    ("/reviews/microtek-solar-pcu-review","Microtek Solar PCU Review","microtek solar pcu review",200,"Easy","C","P2"),
    # Havells 3
    ("/reviews/havells-solar-inverter-review","Havells Solar Inverter Review India 2026","havells solar review",1000,"Medium","C","P1"),
    ("/reviews/havells-enviro-grid-tie-review","Havells Enviro Grid-Tie Review","havells enviro review",250,"Easy","C","P2"),
    ("/reviews/havells-hybrid-inverter-review","Havells Hybrid Inverter Review","havells hybrid review",200,"Easy","C","P2"),
    # GoodWe 4
    ("/reviews/goodwe-solar-inverter-review","GoodWe Solar Inverter Review India 2026","goodwe review",800,"Medium","C","P1"),
    ("/reviews/goodwe-gw5000-ns-review","GoodWe GW5000-NS Review","goodwe gw5000 review",200,"Easy","C","P2"),
    ("/reviews/goodwe-em-series-review","GoodWe EM Series Hybrid Review","goodwe em review",180,"Easy","C","P2"),
    ("/reviews/goodwe-et-series-review","GoodWe ET Series Review","goodwe et review",150,"Easy","C","P2"),
    # Solis 5
    ("/reviews/solis-solar-inverter-review","Solis Solar Inverter Review India 2026","solis inverter review",1500,"Medium","C","P0"),
    ("/reviews/solis-1p5k-4g-review","Solis 1P5K-4G Review","solis 1p5k review",250,"Easy","C","P1"),
    ("/reviews/solis-rhi-3p10k-hves-review","Solis RHI-3P10K-HVES Hybrid Review","solis rhi 3p10k review",200,"Easy","C","P2"),
    ("/reviews/solis-mini-2g-review","Solis Mini 2G Review","solis mini 2g review",180,"Easy","C","P2"),
    ("/reviews/solis-25k-5g-review","Solis 25K-5G Commercial Review","solis 25k review",150,"Easy","C","P2"),
    # Others
    ("/reviews/delta-solar-inverter-review","Delta Solar Inverter Review India","delta solar review",500,"Medium","C","P1"),
    ("/reviews/delta-m6a-review","Delta M6A Review","delta m6a review",150,"Easy","C","P2"),
    ("/reviews/delta-rpi-h5a-review","Delta RPI H5A Review","delta rpi h5a review",120,"Easy","C","P2"),
    ("/reviews/fronius-solar-inverter-review","Fronius Solar Inverter Review India","fronius review india",600,"Medium","C","P1"),
    ("/reviews/fronius-primo-review","Fronius Primo Review","fronius primo review",200,"Easy","C","P2"),
    ("/reviews/fronius-symo-review","Fronius Symo Review","fronius symo review",180,"Easy","C","P2"),
    ("/reviews/sma-solar-inverter-review","SMA Solar Inverter Review India","sma review india",500,"Medium","C","P1"),
    ("/reviews/sma-sunny-boy-review","SMA Sunny Boy Review","sma sunny boy review",250,"Easy","C","P2"),
    ("/reviews/sma-sunny-tripower-review","SMA Sunny Tripower Review","sma tripower review",180,"Easy","C","P2"),
    ("/reviews/huawei-fusionsolar-review","Huawei FusionSolar Review India","huawei fusionsolar review",600,"Medium","C","P1"),
    ("/reviews/huawei-sun2000-l1-review","Huawei SUN2000-L1 Review","huawei sun2000 l1 review",200,"Easy","C","P2"),
    ("/reviews/huawei-sun2000-ktl-m1-review","Huawei SUN2000-KTL-M1 Review","huawei sun2000 ktl review",180,"Easy","C","P2"),
    ("/reviews/enphase-microinverter-review","Enphase Microinverter Review India","enphase microinverter review",400,"Medium","C","P2"),
    ("/reviews/enphase-iq8-review","Enphase IQ8 Review","enphase iq8 review",200,"Easy","C","P2"),
    ("/reviews/solaredge-solar-inverter-review","SolarEdge Solar Inverter Review India","solaredge review",500,"Medium","C","P2"),
    ("/reviews/solaredge-hd-wave-review","SolarEdge HD Wave Review","solaredge hd wave review",200,"Easy","C","P2"),
    ("/reviews/ksolare-solar-inverter-review","K'Solare Solar Inverter Review","ksolare review",200,"Easy","C","P2"),
    ("/reviews/ksolare-hybrid-review","K'Solare Hybrid Review","ksolare hybrid review",100,"Easy","C","P2"),
    ("/reviews/gronsol-solar-inverter-review","Gronsol Solar Inverter Review","gronsol review",100,"Easy","C","P2"),
    ("/reviews/gronsol-hybrid-review","Gronsol Hybrid Review","gronsol hybrid review",80,"Easy","C","P2"),
    ("/reviews/polycab-solar-inverter-review","Polycab Solar Inverter Review","polycab solar review",400,"Medium","C","P2"),
    ("/reviews/polycab-grid-tie-review","Polycab Grid-Tie Inverter Review","polycab grid tie review",150,"Easy","C","P2"),
    ("/reviews/v-guard-solar-inverter-review","V-Guard Solar Inverter Review","v-guard solar review",300,"Easy","C","P2"),
    ("/reviews/v-guard-hybrid-review","V-Guard Hybrid Review","v-guard hybrid review",150,"Easy","C","P2"),
    ("/reviews/sukam-solar-inverter-review","Su-Kam Solar Inverter Review","sukam solar review",500,"Easy","C","P1"),
    ("/reviews/sukam-brainy-eco-review","Su-Kam Brainy Eco Review","sukam brainy review",180,"Easy","C","P2"),
    ("/reviews/sukam-falcon-eco-review","Su-Kam Falcon Eco Review","sukam falcon review",150,"Easy","C","P2"),
    ("/reviews/exide-solar-inverter-review","Exide Solar Inverter Review","exide solar review",400,"Easy","C","P2"),
    ("/reviews/exide-magic-solar-review","Exide Magic Solar Review","exide magic solar review",150,"Easy","C","P2"),
    ("/reviews/waaree-solar-inverter-review","Waaree Solar Inverter Review India","waaree inverter review",800,"Medium","C","P0"),
    ("/reviews/waaree-hybrid-inverter-review","Waaree Hybrid Inverter Review","waaree hybrid review",200,"Easy","C","P1"),
    ("/reviews/waaree-grid-tie-inverter-review","Waaree Grid-Tie Review","waaree grid tie review",180,"Easy","C","P2"),
    ("/reviews/adani-solar-inverter-review","Adani Solar Inverter Review India","adani solar inverter review",600,"Medium","C","P1"),
    ("/reviews/adani-shine-review","Adani Shine Series Review","adani shine review",180,"Easy","C","P2"),
    ("/reviews/utl-solar-inverter-review","UTL Solar Inverter Review India","utl review",1000,"Medium","C","P0"),
    ("/reviews/utl-gamma-plus-review","UTL Gamma Plus Review","utl gamma plus review",400,"Easy","C","P1"),
    ("/reviews/utl-heliac-review","UTL Heliac Review","utl heliac review",250,"Easy","C","P2"),
    ("/reviews/loom-solar-inverter-review","Loom Solar Inverter Review","loom solar review",500,"Medium","C","P1"),
    ("/reviews/loom-fusion-hybrid-review","Loom Fusion Hybrid Review","loom fusion review",180,"Easy","C","P2"),
    ("/reviews/su-vastika-solar-inverter-review","Su-Vastika Solar Inverter Review","suvastika review",200,"Easy","C","P2"),
    ("/reviews/servotech-solar-inverter-review","Servotech Solar Inverter Review","servotech review",150,"Easy","C","P2"),
    ("/reviews/refex-solar-inverter-review","Refex Solar Inverter Review","refex solar review",120,"Easy","C","P2"),
    ("/reviews/jakson-solar-inverter-review","Jakson Solar Inverter Review","jakson solar review",100,"Easy","C","P2"),
    ("/reviews/insolation-solar-inverter-review","Insolation Solar Inverter Review","insolation solar review",80,"Easy","C","P2"),
]
add_sheet("05_Competitor_Reviews", ["URL","Meta Title","Target Keyword","Vol (mo)","KD","Intent","Priority"], reviews, [55, 52, 30, 10, 10, 8, 9])

# ========== 06 GLOSSARY ==========
basic = ["solar-inverter","on-grid-inverter","off-grid-inverter","hybrid-inverter","string-inverter","micro-inverter","mppt","pwm","ac","dc","solar-panel","pv-module","monocrystalline","polycrystalline","bifacial-panel","rooftop-solar","grid-tied","net-metering","gross-metering","net-billing","peak-sun-hours","solar-irradiance","ghi","dni","azimuth","tilt-angle","array","string","combiner-box","junction-box","dc-cable","ac-cable","battery-bank","lithium-ion-battery","lead-acid-battery","inverter-efficiency","cec-efficiency","european-efficiency","nominal-power","peak-power","wattage","kilowatt-hour","capacity-utilization-factor","plant-load-factor","solar-yield","specific-yield","payback-period","lcoe","degradation-rate","solar-warranty"]
tech = ["thd","ip-rating","nema-rating","surge-protection-device","dc-isolator","ac-isolator","anti-islanding","zero-export","reactive-power","apparent-power","real-power","power-factor","transformerless-inverter","galvanic-isolation","ground-fault-detection","arc-fault-detection","rcd","rccb","mcb","modbus","rs485","sunspec","iec-61727","iec-62116","iec-62109","is-17387","ce-marking","rohs","cetl-certification","ul-1741","vde-ar-n","ride-through","low-voltage-ride-through","frt","harmonic-distortion","switching-frequency","silicon-carbide-sic","igbt","mosfet","dc-link-voltage","mppt-tracker-count","string-current","open-circuit-voltage","short-circuit-current","nominal-voltage","self-consumption","battery-soc","battery-dod","cycle-life","round-trip-efficiency"]
finance = ["pm-surya-ghar","mnre","almm-list","capex-model","opex-model","ppa","resco-model","accelerated-depreciation","solar-emi","green-loan","ireda","sbi-solar-finance","gst-on-solar","customs-duty-on-inverters","bcd","perform-achieve-trade","renewable-purchase-obligation","rec-certificate","open-access","wheeling-charge","banking-charge","cross-subsidy-surcharge","discom","cerc","serc","central-financial-assistance","subsidy-disbursement-process","carbon-credit","iso-14001","esg-reporting"]
install = ["site-survey","shading-analysis","structural-load","mounting-structure","rail-mount","ballast-mount","clamps","earthing","lightning-arrester","surge-arrester","cable-tray","conduit","crimping","mc4-connector","commissioning","acceptance-test","as-built-drawing","ceig-approval","electrical-inspector","fire-noc"]

glossary_rows = []
def gloss_row(term, cat, prio="P1"):
    name = term.replace('-', ' ').title()
    return [f"/glossary/{term}", f"What is {name}? — Solar Glossary | Qbits", f"what is {term.replace('-',' ')}",
            300 if cat=="Basic" else 200, "Easy", "I", prio, "Not Started", "Glossary", 750,
            "DefinedTerm+FAQ", "3 related glossary + 1 product", f"{cat} term"]
for t in basic: glossary_rows.append(gloss_row(t, "Basic", "P0" if t in ["mppt","hybrid-inverter","on-grid-inverter","net-metering","solar-inverter"] else "P1"))
for t in tech: glossary_rows.append(gloss_row(t, "Technical"))
for t in finance: glossary_rows.append(gloss_row(t, "Finance/Policy", "P0" if t in ["pm-surya-ghar","almm-list","mnre","discom"] else "P1"))
for t in install: glossary_rows.append(gloss_row(t, "Installation", "P2"))

add_sheet("06_Glossary_Pages", headers, glossary_rows, [38, 50, 30, 10, 10, 8, 9, 14, 12, 11, 22, 30, 30])

# ========== 07 CITY PAGES ==========
tier1 = ["mumbai","delhi","bangalore","chennai","hyderabad","kolkata","pune","ahmedabad"]
tier2 = ["surat","jaipur","lucknow","kanpur","nagpur","indore","bhopal","patna","vadodara","ludhiana","agra","nashik","faridabad","meerut","rajkot","varanasi","srinagar","aurangabad","dhanbad","amritsar","navi-mumbai","allahabad","ranchi","howrah","coimbatore","jabalpur","gwalior","vijayawada","jodhpur","madurai","raipur","kota","chandigarh","guwahati","solapur","hubli-dharwad","bareilly","moradabad","mysore","gurgaon","aligarh","jalandhar"]
tier3 = ["tiruppur","salem","erode","vellore","trichy","kanchipuram","thanjavur","nellore","guntur","tirupati","visakhapatnam","warangal","karimnagar","nizamabad","mangalore","belgaum","gulbarga","davangere","bellary","shimoga","tumkur","udupi","mysuru","kochi","thiruvananthapuram","kozhikode","thrissur","kollam","palakkad","alappuzha","kannur","kottayam","gandhinagar","bhavnagar","jamnagar","junagadh","anand","bharuch","navsari","mehsana","valsad","nadiad","porbandar","noida","ghaziabad","gorakhpur","saharanpur","firozabad","jhansi","muzaffarnagar","mathura","rampur","shahjahanpur","faizabad","ayodhya","etawah","mirzapur","bulandshahr","sambhal","amroha","hardoi","sitapur","bahraich","barabanki","unnao","jaunpur","azamgarh","dehradun","haldwani","rishikesh","roorkee","haridwar","ajmer","udaipur","bikaner","alwar","bhilwara","sikar","pali","sri-ganganagar","kishangarh","beawar","hanumangarh","ujjain","sagar","dewas","satna","ratlam","rewa","murwara","burhanpur","khandwa","bhilai","bilaspur","korba","durg","rajnandgaon","panaji","silvassa","daman","puducherry","port-blair"]

city_rows = []
def city_row(c, tier):
    name = c.replace('-',' ').title()
    vol = {"T1":2500, "T2":900, "T3":350}[tier]
    prio = {"T1":"P0", "T2":"P1", "T3":"P2"}[tier]
    return [f"/solar-inverter/{c}", f"Solar Inverter in {name} | Best Brands, Price, Installation",
            f"solar inverter {c.replace('-',' ')}", vol, "Medium", "L", prio, "Not Started",
            "Location", 1800, "LocalBusiness+FAQ+Breadcrumb", "State page + 3 product + subsidy + nearest cities",
            f"{tier} city; local DISCOM, irradiance, installer, ROI math, testimonial"]
for c in tier1: city_rows.append(city_row(c, "T1"))
for c in tier2: city_rows.append(city_row(c, "T2"))
for c in tier3[:100]: city_rows.append(city_row(c, "T3"))

add_sheet("07_City_Pages", headers, city_rows, [38, 55, 30, 10, 10, 8, 9, 14, 12, 11, 24, 38, 50])

# ========== 08 STATE PAGES ==========
states = ["gujarat","maharashtra","tamil-nadu","karnataka","uttar-pradesh","rajasthan","madhya-pradesh","andhra-pradesh","telangana","kerala","west-bengal","punjab","haryana","delhi","odisha","bihar","jharkhand","chhattisgarh","assam","uttarakhand","himachal-pradesh","jammu-kashmir","goa","tripura","meghalaya","manipur","nagaland","arunachal-pradesh"]
state_rows = []
priority_states = ["gujarat","maharashtra","tamil-nadu","uttar-pradesh","karnataka","rajasthan","delhi","andhra-pradesh","telangana","madhya-pradesh"]
for s in states:
    name = s.replace('-',' ').title()
    vol = 1500 if s in priority_states else 600
    prio = "P0" if s in priority_states else "P1"
    state_rows.append([f"/solar-inverter/{s}", f"Solar Inverter in {name} | Subsidy, Price, Installer",
        f"solar inverter {s.replace('-',' ')}", vol, "Medium", "L", prio, "Not Started",
        "State Hub", 2000, "LocalBusiness+FAQ+Place", "All state cities + state subsidy + scheme hub",
        "State irradiance, DISCOM, scheme, top 5 cities, ROI"])
add_sheet("08_State_Pages", headers, state_rows, [38, 55, 30, 10, 10, 8, 9, 14, 12, 11, 24, 38, 50])

# ========== 09 BLOG PAGES ==========
blog = [
    # How-to & buying guides (40)
    ("/blog/how-to-choose-solar-inverter-for-home-india","How to Choose a Solar Inverter for Your Home in India (2026 Guide)","how to choose solar inverter",2500,"Hard","I","P0","BG","Pillar buying guide"),
    ("/blog/how-to-size-solar-inverter","How to Size a Solar Inverter — Step-by-Step","solar inverter sizing",1500,"Medium","I","P0","BG","Sizing pillar"),
    ("/blog/how-to-read-inverter-datasheet","How to Read a Solar Inverter Datasheet","read inverter datasheet",400,"Easy","I","P1","BG","Technical depth"),
    ("/blog/inverter-buying-guide-india","Solar Inverter Buying Guide for India 2026","solar inverter buying guide",1500,"Medium","I","P0","BG","Pillar"),
    ("/blog/how-many-solar-panels-need","How Many Solar Panels Do I Need for My Home?","how many solar panels need",4000,"Hard","I","P0","BG","Sizing"),
    ("/blog/how-does-solar-inverter-work","How Does a Solar Inverter Work? (With Diagram)","how does solar inverter work",6000,"Hard","I","P0","BG","Foundational"),
    ("/blog/on-grid-vs-hybrid-explained","On-Grid vs Hybrid Inverter — Which One Should You Buy?","on grid vs hybrid",2200,"Hard","I/C","P0","BG","Decision pillar"),
    ("/blog/best-solar-inverter-brands-explained","Top 10 Solar Inverter Brands in India Explained","best solar inverter brands explained",1500,"Hard","C","P0","BG","Brand authority"),
    ("/blog/solar-inverter-life-expectancy","How Long Does a Solar Inverter Last?","solar inverter life",1200,"Medium","I","P1","BG","Longevity"),
    ("/blog/inverter-maintenance-tips","10 Solar Inverter Maintenance Tips","solar inverter maintenance",1500,"Medium","I","P1","BG","Care"),
    # Subsidy & policy (30) — main 10 here
    ("/blog/pm-surya-ghar-explained","PM Surya Ghar Muft Bijli Yojana — Complete Guide 2026","pm surya ghar yojana",200000,"Very Hard","I","P0","SUB","Top traffic driver"),
    ("/blog/pm-surya-ghar-eligibility","PM Surya Ghar Eligibility — Who Can Apply?","pm surya ghar eligibility",20000,"Hard","I","P0","SUB","High-intent"),
    ("/blog/pm-surya-ghar-subsidy-calculator","PM Surya Ghar Subsidy Calculator — Find Your Amount","pm surya ghar subsidy amount",15000,"Hard","T","P0","SUB","Tool/blog hybrid"),
    ("/blog/almm-list-explained","ALMM List Phase III — Complete Guide for Buyers","almm list",4000,"Hard","I","P0","SUB","Authority topic"),
    ("/blog/mnre-rooftop-scheme","MNRE Rooftop Solar Scheme — How It Works","mnre rooftop scheme",3500,"Hard","I","P0","SUB","Govt scheme"),
    ("/blog/kusum-scheme-explained","KUSUM Yojana — Agricultural Solar Subsidy Guide","kusum yojana",30000,"Very Hard","I","P0","SUB","Ag farmers"),
    ("/blog/net-metering-explained","Net Metering in India — Complete Guide 2026","net metering india",4000,"Hard","I","P0","SUB","Foundational"),
    ("/blog/gross-metering-vs-net-metering","Gross Metering vs Net Metering Explained","gross vs net metering",1200,"Medium","I","P1","SUB","Decision"),
    ("/blog/gst-on-solar-2026","GST on Solar Inverters & Panels 2026","gst on solar inverter",1500,"Medium","I","P1","SUB","Buyer Q"),
    ("/blog/discom-empanelled-vendor","What is an Empanelled Vendor? Why It Matters","empanelled vendor solar",600,"Easy","I","P1","SUB","Process"),
    # Industry analysis (20) — main 10 here
    ("/blog/india-rooftop-solar-index-2026","India Rooftop Solar Index 2026 — Annual Report","india rooftop solar 2026",1500,"Medium","I","P1","IA","Annual link bait"),
    ("/blog/solar-inverter-market-india","Solar Inverter Market in India 2026","solar inverter market india",1000,"Medium","I","P1","IA","Industry view"),
    ("/blog/ai-in-solar-inverters","How AI Is Changing Solar Inverter Performance","ai solar inverter india",400,"Easy","I","P0","IA","Qbits USP narrative"),
    ("/blog/atmanirbhar-bharat-solar","Atmanirbhar Bharat & Indian Solar Manufacturing","atmanirbhar bharat solar",300,"Easy","I","P0","IA","Made-in-India angle"),
    ("/blog/discom-net-metering-comparison","India's 28 DISCOMs — Net Metering Compared","discom net metering compare",400,"Easy","I","P1","IA","Authority depth"),
    ("/blog/solar-vs-diesel-generator","Solar vs Diesel Generator — Cost & Payback","solar vs diesel generator",1200,"Medium","C","P0","IA","Industrial framing"),
    ("/blog/lithium-vs-lead-acid-solar","Lithium vs Lead-Acid Solar Battery — 2026","lithium vs lead acid solar",800,"Medium","C","P1","IA","Battery"),
    ("/blog/string-vs-microinverter-india","String vs Microinverter — Indian Roof Reality","string vs microinverter india",500,"Medium","I","P1","IA","Tech buyer"),
    ("/blog/solar-panel-degradation-india","Solar Panel Degradation in Indian Climate","solar panel degradation",600,"Medium","I","P1","IA","Long-term"),
    ("/blog/inverter-warranty-explained-india","Solar Inverter Warranty in India — Honest Look","inverter warranty india",500,"Medium","I","P0","IA","Qbits 12-yr USP"),
    # Case studies (40) — main 10 here
    ("/blog/case-study-pune-3kw-home","Case Study — 3kW Solar on a Pune Home","pune solar case study",200,"Easy","I","P1","CS","Residential"),
    ("/blog/case-study-tirupur-textile-mill","Case Study — 100kW Solar for a Tirupur Textile Mill","tirupur textile solar",150,"Easy","I","P0","CS","Industrial flagship"),
    ("/blog/case-study-jaipur-villa-10kw","Case Study — 10kW Solar on a Jaipur Villa","jaipur solar villa",150,"Easy","I","P1","CS","Premium residential"),
    ("/blog/case-study-bangalore-hospital-50kw","Case Study — 50kW Solar at a Bangalore Hospital","bangalore hospital solar",120,"Easy","I","P1","CS","Healthcare"),
    ("/blog/case-study-coimbatore-school","Case Study — 25kW Solar at a Coimbatore School","coimbatore school solar",100,"Easy","I","P1","CS","School"),
    ("/blog/case-study-rajkot-cold-storage","Case Study — 40kW Solar at a Rajkot Cold Storage","rajkot cold storage solar",100,"Easy","I","P1","CS","Cold storage"),
    ("/blog/case-study-lucknow-3-bhk","Case Study — 5kW on a 3 BHK in Lucknow","lucknow solar 3 bhk",100,"Easy","I","P1","CS","UP residential"),
    ("/blog/case-study-dairy-farm-punjab","Case Study — Solar on a Punjab Dairy Farm","punjab dairy farm solar",100,"Easy","I","P2","CS","Ag"),
    ("/blog/case-study-noida-apartment","Case Study — Solar on a Noida Apartment Complex","noida apartment solar",120,"Easy","I","P2","CS","RWA"),
    ("/blog/case-study-chennai-it-park","Case Study — 200kW Solar at a Chennai IT Park","chennai it park solar",100,"Easy","I","P2","CS","Commercial"),
    # Founder POV / thought leadership (20) — main 10
    ("/blog/founder-on-service-sla-crisis","Why India's Solar Industry Must Solve the Service SLA Crisis","solar service sla india",100,"Easy","I","P0","TL","Founder POV"),
    ("/blog/founder-on-india-grid-tuning","How We Tune Inverters for the Indian Grid","ai india grid solar inverter",200,"Easy","I","P0","TL","USP narrative"),
    ("/blog/founder-on-installer-economics","The Unit Economics of an Indian Solar Installer","solar installer economics india",200,"Easy","I","P1","TL","B2B influence"),
    ("/blog/founder-on-channel-protection","Why Qbits Will Never Be Sold on Amazon","solar inverter channel protection",100,"Easy","I","P1","TL","Trade-press hook"),
    ("/blog/founder-on-data-sovereignty","Where Your Solar Monitoring Data Should Live","solar data sovereignty india",100,"Easy","I","P1","TL","Made-for-India"),
    ("/blog/founder-on-warranty","Why a 12-Year Warranty Is the Bare Minimum","solar inverter warranty india",200,"Easy","I","P1","TL","USP"),
    ("/blog/founder-on-pm-surya-ghar","PM Surya Ghar — What's Working, What Isn't","pm surya ghar opinion",200,"Easy","I","P1","TL","Policy POV"),
    ("/blog/founder-on-hybrid-vs-grid","Why Most Indian Buyers Should Choose Hybrid","why hybrid solar inverter india",300,"Easy","I","P1","TL","Hybrid narrative"),
    ("/blog/founder-on-tier3-solar","Tier-3 India Is the Real Solar Story","tier 3 solar india",100,"Easy","I","P2","TL","Geographic POV"),
    ("/blog/founder-on-discom-paperwork","Fixing India's DISCOM Net-Metering Paperwork","discom paperwork solar",80,"Easy","I","P2","TL","Process pain"),
    # Seasonal / news-jacking (30) — main 10
    ("/blog/monsoon-solar-prep","Prepping Your Solar System for Indian Monsoon","monsoon solar prep",800,"Medium","I","P0","SEAS","June-Sept relevant"),
    ("/blog/summer-inverter-derating","Why Your Solar Inverter Produces Less in Summer Heat","inverter derating summer india",400,"Easy","I","P0","SEAS","Apr-May"),
    ("/blog/winter-solar-india","Solar Performance in Indian Winters","solar performance winter india",300,"Easy","I","P1","SEAS","Dec-Feb"),
    ("/blog/budget-2026-solar-impact","Union Budget 2026 — What It Means for Solar","budget 2026 solar",1500,"Medium","I","P0","SEAS","Feb annual"),
    ("/blog/pm-surya-ghar-q1-2026-update","PM Surya Ghar — Q1 2026 Update","pm surya ghar 2026 update",2500,"Hard","I","P0","SEAS","Quarterly"),
    ("/blog/almm-phase-3-update","ALMM Phase III Update 2026","almm phase 3 update",800,"Medium","I","P1","SEAS","Policy news"),
    ("/blog/intersolar-2026-recap","Intersolar India 2026 — Top Trends Recap","intersolar 2026",500,"Medium","I","P2","SEAS","Event"),
    ("/blog/rei-2026-recap","Renewable Energy India Expo 2026 — Recap","rei 2026 recap",400,"Medium","I","P2","SEAS","Event"),
    ("/blog/electricity-tariff-hike-2026","Electricity Tariff Hikes in 2026 — Why Solar Now","electricity tariff hike 2026",1500,"Medium","I","P0","SEAS","Newsjack"),
    ("/blog/g20-india-renewables","India's G20 Climate Commitments & Rooftop Solar","g20 india renewables",300,"Easy","I","P2","SEAS","Policy"),
]
# Pad blog list (note in README that full editorial calendar = 180)
blog_headers = ["URL","Meta Title","Target Keyword","Vol (mo)","KD","Intent","Priority","Category","Notes"]
add_sheet("09_Blog_Pages", blog_headers, blog, [50, 55, 30, 10, 10, 8, 9, 9, 40])

# ========== 10 SUBSIDY PAGES ==========
sub_rows = []
priority_sub = ["gujarat","maharashtra","tamil-nadu","uttar-pradesh","karnataka","rajasthan","delhi","kerala","andhra-pradesh","telangana"]
for s in states:
    name = s.replace('-',' ').title()
    vol = 3000 if s in priority_sub else 800
    prio = "P0" if s in priority_sub else "P1"
    sub_rows.append([f"/subsidy/{s}", f"Solar Subsidy in {name} 2026 — Amount, Eligibility, Apply",
        f"solar subsidy {s.replace('-',' ')}", vol, "Medium", "I", prio, "Not Started",
        "Subsidy Hub", 2200, "HowTo+FAQ+Article", "State page + scheme hub + city pages",
        "State-specific subsidy + DISCOM steps + case study"])
# Scheme hubs
sub_rows.append(["/subsidy/pm-surya-ghar","PM Surya Ghar Yojana 2026 — Complete Guide","pm surya ghar",200000,"Very Hard","I","P0","Not Started","Scheme Hub",3500,"HowTo+FAQ+Article","All state subsidy","Master scheme hub"])
sub_rows.append(["/subsidy/kusum","KUSUM Yojana — Agricultural Solar Subsidy","kusum scheme",30000,"Very Hard","I","P0","Not Started","Scheme Hub",2500,"HowTo+FAQ","State subsidy + ag use-case","Farmer pumps"])
sub_rows.append(["/subsidy/mnre-rooftop","MNRE Rooftop Solar Scheme 2026","mnre rooftop scheme",3500,"Hard","I","P0","Not Started","Scheme Hub",2000,"HowTo+FAQ","PM Surya Ghar","Parent scheme"])
sub_rows.append(["/subsidy/state-comparison","State-wise Solar Subsidy Comparison 2026","state solar subsidy compare",800,"Medium","I","P0","Not Started","Comparison Hub",2200,"ItemList+FAQ","All state subsidy","Comparison authority"])
add_sheet("10_Subsidy_Pages", headers, sub_rows, [40, 55, 30, 10, 10, 8, 9, 14, 12, 11, 22, 30, 40])

# ========== 11 CALCULATORS ==========
calcs = [
    ("/calculator/system-sizing","Solar System Sizing Calculator India — Free","solar system sizing calculator",2000,"Hard","C","P0","Master sizing tool"),
    ("/calculator/payback","Solar Payback Calculator India 2026 — Free","solar payback calculator india",2500,"Hard","C","P0","Embeddable link magnet"),
    ("/calculator/subsidy","PM Surya Ghar Subsidy Calculator — Find Your Amount","pm surya ghar calculator",15000,"Hard","T","P0","High-volume tool"),
    ("/calculator/emi","Solar EMI Calculator India","solar emi calculator",1800,"Medium","T","P0","Finance angle"),
    ("/calculator/roi-commercial","Commercial Solar ROI Calculator India","commercial solar roi calculator",800,"Medium","C","P1","C&I lead magnet"),
    ("/calculator/savings","Solar Bill Savings Calculator","solar bill savings calculator",2000,"Medium","T","P0","BOFU"),
    ("/calculator/co2-offset","Solar CO2 Offset Calculator India","solar co2 calculator",400,"Easy","I","P1","ESG/CSR angle"),
    ("/calculator/battery-sizing","Solar Battery Sizing Calculator","solar battery sizing calculator",600,"Medium","C","P1","Hybrid buyer"),
    ("/calculator/inverter-sizing","Solar Inverter Sizing Calculator","solar inverter sizing calculator",1500,"Medium","C","P0","Inverter-specific"),
    ("/tools/load-estimator","Home Load Estimator for Solar","home load calculator solar",800,"Medium","C","P1","Pre-sizing"),
    ("/tools/bill-analyzer","Upload Electricity Bill — Get Solar Quote","electricity bill solar analyzer",300,"Easy","T","P0","Lead capture via WhatsApp upload"),
    ("/tools/installer-finder","Find Qbits-Certified Solar Installer Near You","solar installer near me",5000,"Hard","L","P0","Pincode tool"),
]
add_sheet("11_Calculator_Tools", ["URL","Meta Title","Target Keyword","Vol (mo)","KD","Intent","Priority","Notes"], calcs, [38, 55, 32, 10, 10, 8, 9, 50])

# ========== 12 USE CASE / VERTICALS ==========
verticals = [
    ("/solar-for-home","Solar for Home India — Complete Guide & Quote","solar for home india",12000,"Very Hard","C","P0","Residential master"),
    ("/solar-for-apartment","Solar for Apartment & Flat in India","solar for apartment india",1500,"Medium","C","P0","Urban dense"),
    ("/solar-for-villa","Solar for Villa & Bungalow in India","solar for villa india",800,"Medium","C","P1","Premium res"),
    ("/solar-for-housing-society","Solar for Housing Society / RWA","solar for housing society",600,"Medium","C","P1","RWA / society"),
    ("/solar-for-shop","Solar for Shop / Retail Outlet","solar for shop india",2000,"Medium","C","P0","SMB"),
    ("/solar-for-office","Solar for Office Space","solar for office",1500,"Medium","C","P0","SMB"),
    ("/solar-for-factory","Solar for Factory & MSME","solar for factory india",1500,"Medium","C","P0","C&I anchor"),
    ("/solar-for-textile-mill","Solar for Textile Mill (Tirupur/Surat)","solar for textile mill",300,"Easy","C","P0","Tier-3 industrial"),
    ("/solar-for-pharma","Solar for Pharma Manufacturing","solar for pharma",250,"Easy","C","P1","Vertical"),
    ("/solar-for-cold-storage","Solar for Cold Storage","solar for cold storage",300,"Easy","C","P0","Niche high-value"),
    ("/solar-for-petrol-pump","Solar for Petrol Pump","solar for petrol pump",600,"Medium","C","P0","Vertical-specific"),
    ("/solar-for-water-pump","Solar Water Pump under KUSUM","solar water pump india",2500,"Hard","C","P0","KUSUM scheme"),
    ("/solar-for-atta-chakki","Solar for Atta Chakki / Flour Mill","solar for atta chakki",400,"Easy","C","P1","Tier-2/3"),
    ("/solar-for-poultry-farm","Solar for Poultry Farm","solar for poultry farm",400,"Easy","C","P1","Ag"),
    ("/solar-for-dairy","Solar for Dairy Farm","solar for dairy",300,"Easy","C","P1","Ag"),
    ("/solar-for-school","Solar for School (CSR / Institutional)","solar for school",800,"Medium","C","P0","Institutional"),
    ("/solar-for-hospital","Solar for Hospital","solar for hospital",500,"Medium","C","P0","Critical load"),
    ("/solar-for-hotel","Solar for Hotel / Resort","solar for hotel",400,"Easy","C","P1","Hospitality"),
    ("/solar-for-mall","Solar for Shopping Mall","solar for mall",250,"Easy","C","P2","Commercial"),
    ("/solar-for-warehouse","Solar for Warehouse / Logistics","solar for warehouse",500,"Medium","C","P1","Logistics"),
    ("/solar-for-printing-press","Solar for Printing Press","solar for printing press",150,"Easy","C","P2","Niche"),
    ("/solar-for-ro-plant","Solar for RO / Water Purification Plant","solar for ro plant",350,"Easy","C","P1","Vertical"),
    ("/solar-for-restaurant","Solar for Restaurant / Cloud Kitchen","solar for restaurant",250,"Easy","C","P2","F&B"),
    ("/solar-for-clinic","Solar for Clinic / Diagnostic Center","solar for clinic",200,"Easy","C","P2","Healthcare SMB"),
    ("/solar-for-gym","Solar for Gym / Fitness Center","solar for gym",150,"Easy","C","P2","SMB"),
    ("/solar-for-telecom-tower","Solar for Telecom Tower","solar for telecom tower",250,"Easy","C","P2","B2B niche"),
    ("/solar-for-ice-factory","Solar for Ice Factory","solar for ice factory",150,"Easy","C","P2","Niche"),
    ("/solar-for-saw-mill","Solar for Saw Mill","solar for saw mill",100,"Easy","C","P2","Niche"),
    ("/solar-for-ev-charging","Solar + EV Charging at Home","solar ev charging india",500,"Medium","C","P0","Future-proof"),
    ("/solar-for-flour-mill","Solar for Flour Mill","solar for flour mill",200,"Easy","C","P2","Niche"),
]
add_sheet("12_UseCase_Verticals", ["URL","Meta Title","Target Keyword","Vol (mo)","KD","Intent","Priority","Notes"], verticals, [38, 55, 32, 10, 10, 8, 9, 50])

# ========== 13 INSTALLER HUB ==========
installer = [
    ("/installer","Qbits Installer Hub — Tools, Training, Leads","solar installer hub india",500,"Medium","I","P0","Hub home"),
    ("/installer/become-a-partner","Become a Qbits Authorized Installer","solar installer partnership india",400,"Medium","T","P0","Partner sign-up"),
    ("/installer/certification","Qbits Installer Certification — Bronze/Silver/Gold","solar installer certification india",200,"Easy","C","P1","Tiered program"),
    ("/installer/leads","Lead-Share Program for Qbits Installers","solar installer leads",300,"Easy","T","P0","Magnet for installers"),
    ("/installer/marketing-kit","Free Marketing Kit for Solar Installers","solar installer marketing kit",150,"Easy","I","P1","Asset gift"),
    ("/installer/sla-tracker","Live RMA & Spare-Parts SLA Tracker","solar spares sla",80,"Easy","I","P1","Transparency"),
    ("/installer/community","Qbits Installer WhatsApp Community","solar installer community india",100,"Easy","I","P1","Community"),
    ("/installer/case-studies","Solar Installer Success Stories","solar installer case study",150,"Easy","I","P1","Social proof"),
    # 20 training modules
    *[(f"/installer/training/module-{i:02d}", f"Installer Training Module {i:02d}", f"solar installer training module {i}", 100, "Easy", "I", "P1", f"Module {i}") for i in range(1, 21)],
    # 28 DISCOM paperwork pages
    *[(f"/installer/discom-paperwork/{s}", f"DISCOM Paperwork Guide — {s.replace('-',' ').title()}", f"discom net metering {s.replace('-',' ')}", 200, "Easy", "I", "P1", f"{s.title()} forms") for s in states[:28]],
    # 4 misc
    ("/installer/discom-paperwork-master","India's Master DISCOM Net Metering Paperwork Guide","discom paperwork india",400,"Medium","I","P1","Master guide"),
    ("/installer/sizing-tools","Installer Sizing & Design Tools","solar installer sizing tool",200,"Easy","I","P1","Tool gift"),
    ("/installer/commissioning-checklist","Solar Commissioning Checklist (PDF + Interactive)","solar commissioning checklist",300,"Easy","I","P1","Field tool"),
    ("/installer/troubleshooting-guide","Field Troubleshooting Guide for Installers","solar troubleshooting guide",250,"Easy","I","P1","Field service"),
]
add_sheet("13_Installer_Hub", ["URL","Meta Title","Target Keyword","Vol (mo)","KD","Intent","Priority","Notes"], installer, [50, 55, 32, 10, 10, 8, 9, 50])

# ========== 14 ERROR CODES ==========
err_codes = ["isolation-fault","grid-overvoltage","grid-undervoltage","grid-frequency-error","no-power-output","display-blank","overheating","fan-noise","wifi-not-connecting","app-not-syncing","dc-overvoltage","dc-undervoltage","ground-fault","arc-fault","reverse-polarity","mppt-error","string-mismatch","insulation-resistance-low","battery-overvoltage","battery-undervoltage"]
sku_families = [("qh5","QH5"),("qh10t","QH10T"),("q5s","Q5S"),("q25t","Q25T")]
err_rows = []
for sku, label in sku_families:
    for code in err_codes:
        nice = code.replace('-',' ').title()
        err_rows.append([
            f"/support/error-code/{sku}-{code}",
            f"{label} — {nice} | Solution & Field Fix",
            f"{label} {nice.lower()} error",
            120, "Easy", "I", "P1", "Not Started", "Support",
            500, "FAQPage+TroubleshootingGuide", "Product page + glossary",
            f"{label} {nice}: cause, fix, when to escalate"
        ])
add_sheet("14_Error_Codes", headers, err_rows, [40, 55, 30, 10, 10, 8, 9, 14, 12, 11, 26, 28, 40])

# ========== 15 HINDI CONTENT ==========
hindi = [
    ("/hi/","Qbits Energy — Bharat ke liye AI Solar Inverter","qbits hindi",80,"Easy","I","P0","Hindi home"),
    ("/hi/solar-inverter-kya-hai","Solar Inverter Kya Hai? Hindi me Samjhe","solar inverter kya hai",4000,"Medium","I","P0","Top Hindi info"),
    ("/hi/mppt-kya-hota-hai","MPPT Kya Hota Hai? Hindi Guide","mppt kya hota hai",2500,"Easy","I","P0","Low-comp Hindi"),
    ("/hi/on-grid-vs-off-grid","On Grid aur Off Grid me kya antar","on grid off grid antar",1500,"Easy","I","P0","Hindi info"),
    ("/hi/ghar-ke-liye-solar-inverter","Ghar ke liye Solar Inverter kaise chunein","ghar ke liye solar inverter",3000,"Medium","C","P0","Hindi BOFU"),
    ("/hi/pm-surya-ghar-yojana-hindi","PM Surya Ghar Yojana — Hindi me Sampoorn Jaankari","pm surya ghar hindi",30000,"Hard","I","P0","Top Hindi traffic"),
    ("/hi/pm-surya-ghar-kaise-apply-karein","PM Surya Ghar Apply kaise karein — Step by Step","pm surya ghar apply hindi",15000,"Hard","T","P0","Hindi how-to"),
    ("/hi/solar-subsidy-kaise-milegi","Solar Subsidy kaise milegi — Hindi Guide","solar subsidy kaise milegi",4000,"Medium","I","P0","High Hindi intent"),
    ("/hi/5kw-solar-inverter-price-hindi","5kW Solar Inverter Price Hindi — Full Guide","5kw solar inverter price hindi",2000,"Medium","C","P0","Hindi commercial"),
    ("/hi/3kw-solar-inverter-price-hindi","3kW Solar Inverter Price Hindi","3kw solar inverter price hindi",1500,"Medium","C","P0","Hindi commercial"),
    ("/hi/hybrid-solar-inverter-hindi","Hybrid Solar Inverter Kya Hai — Hindi","hybrid solar inverter hindi",1200,"Easy","I","P1","Hindi info"),
    ("/hi/solar-panel-kaise-lagaye","Ghar me Solar Panel kaise lagaye","ghar me solar panel kaise lagaye",2500,"Medium","I","P0","Hindi how-to"),
    ("/hi/bijli-bill-kaise-kam-karein","Bijli Bill kaise kam karein — Solar Solution","bijli bill kam karein",5000,"Medium","C","P0","Hindi top of funnel"),
    ("/hi/atta-chakki-solar-inverter","Atta Chakki ke liye Solar Inverter","atta chakki solar hindi",500,"Easy","C","P0","Tier-3 Hindi vertical"),
    ("/hi/pani-pump-solar-hindi","Pani Pump ke liye Solar — KUSUM Yojana","kusum yojana hindi",8000,"Medium","I","P0","Hindi farmer"),
    ("/hi/solar-inverter-installation-cost","Solar Inverter Installation Cost Hindi","solar inverter installation cost hindi",800,"Easy","C","P1","Hindi pricing"),
    ("/hi/solar-inverter-warranty","Solar Inverter Warranty kitne saal","solar inverter warranty hindi",500,"Easy","I","P1","Hindi trust"),
    ("/hi/solar-inverter-chalu-nahi-ho-raha","Solar Inverter Chalu Nahi Ho Raha — Solution","solar inverter band ho gaya",1500,"Easy","I","P0","Hindi support"),
    ("/hi/solar-inverter-aur-ups-antar","Solar Inverter aur UPS me kya antar","solar inverter ups antar",1000,"Easy","I","P0","Hindi info"),
    ("/hi/anti-islanding-hindi","Anti Islanding Kya Hai — Hindi","anti islanding hindi",300,"Easy","I","P1","Hindi technical"),
    *[(f"/hi/solar-inverter-{c}", f"{c.replace('-',' ').title()} me Solar Inverter — Hindi", f"solar inverter {c.replace('-',' ')} hindi", 500, "Easy", "L", "P1", f"Hindi city: {c.title()}") for c in ["lucknow","kanpur","varanasi","agra","meerut","gorakhpur","allahabad","ghaziabad","noida","faridabad","jaipur","jodhpur","udaipur","kota","bhopal","indore","jabalpur","gwalior","patna","ranchi"]],
    ("/hi/solar-inverter-uttar-pradesh","Uttar Pradesh me Solar Inverter — Hindi Guide","solar inverter up hindi",2500,"Medium","L","P0","UP state Hindi"),
    ("/hi/solar-inverter-bihar","Bihar me Solar Inverter — Hindi","solar inverter bihar hindi",1500,"Medium","L","P0","Bihar state Hindi"),
    ("/hi/solar-inverter-madhya-pradesh","Madhya Pradesh me Solar Inverter — Hindi","solar inverter mp hindi",1200,"Medium","L","P0","MP state Hindi"),
    ("/hi/solar-inverter-rajasthan","Rajasthan me Solar Inverter — Hindi","solar inverter rajasthan hindi",1500,"Medium","L","P0","Rajasthan state Hindi"),
    ("/hi/whatsapp-solar-alerts","Solar Inverter ka WhatsApp Alert — Hindi","whatsapp solar alert hindi",200,"Easy","C","P0","USP in Hindi"),
    ("/hi/best-solar-inverter-india-hindi","Bharat ka Sabse Achha Solar Inverter 2026","best solar inverter india hindi",2000,"Medium","C","P0","Hindi BOFU pillar"),
    ("/hi/luminous-vs-qbits-hindi","Qbits vs Luminous Solar Inverter — Hindi","qbits vs luminous hindi",200,"Easy","C","P0","Hindi comparison"),
    ("/hi/growatt-vs-qbits-hindi","Qbits vs Growatt — Hindi Comparison","qbits vs growatt hindi",150,"Easy","C","P0","Hindi comparison"),
    ("/hi/microtek-vs-qbits-hindi","Qbits vs Microtek — Hindi","qbits vs microtek hindi",150,"Easy","C","P0","Hindi comparison"),
    ("/hi/inverter-overheating-solution","Solar Inverter Garam Ho Raha — Solution Hindi","solar inverter garam hindi",500,"Easy","I","P1","Hindi support"),
]
add_sheet("15_Hindi_Content", ["URL","Meta Title","Target Keyword","Vol (mo)","KD","Intent","Priority","Notes"], hindi, [50, 55, 32, 10, 10, 8, 9, 50])

# ========== 16 GUJARATI CONTENT ==========
guj = [
    ("/gu/","Qbits Energy — Gujarat mate AI Solar Inverter","qbits gujarati",60,"Easy","I","P0","Gujarati home"),
    ("/gu/solar-inverter-shu-che","Solar Inverter Shu Che? Gujarati Guide","solar inverter gujarati",400,"Easy","I","P0","Gujarati info"),
    ("/gu/ghar-mate-solar-inverter","Ghar mate Solar Inverter Kevi Rite Pasand Karvo","ghar mate solar inverter",500,"Easy","C","P0","Gujarati BOFU"),
    ("/gu/surya-gujarat-yojana","Surya Gujarat Yojana — Sampoorn Margdarshan","surya gujarat yojana",2000,"Medium","I","P0","Top Gujarati traffic"),
    ("/gu/pm-surya-ghar-gujarati","PM Surya Ghar Yojana — Gujarati","pm surya ghar gujarati",1500,"Medium","I","P0","Scheme Gujarati"),
    ("/gu/solar-subsidy-gujarat","Solar Subsidy Gujarat — Application Process","solar subsidy gujarat gujarati",1200,"Medium","I","P0","Subsidy in Guj"),
    ("/gu/5kw-solar-inverter-bhav","5kW Solar Inverter no Bhav — Gujarati","5kw solar inverter bhav",400,"Easy","C","P0","Gujarati commercial"),
    ("/gu/3kw-solar-inverter-bhav","3kW Solar Inverter no Bhav — Gujarati","3kw solar inverter bhav",300,"Easy","C","P0","Gujarati commercial"),
    ("/gu/hybrid-solar-inverter","Hybrid Solar Inverter — Gujarati Margdarshan","hybrid solar inverter gujarati",250,"Easy","I","P1","Gujarati info"),
    ("/gu/on-grid-off-grid","On-Grid ane Off-Grid Vache Sherufero","on grid off grid gujarati",200,"Easy","I","P1","Gujarati info"),
    ("/gu/bijli-bill-kam","Bijli Bill Kam Karva ni Rito","bijli bill kam gujarati",800,"Easy","C","P0","Gujarati top funnel"),
    ("/gu/solar-installation-kharcho","Solar Installation no Kharcho — Gujarati","solar installation kharcho",500,"Easy","C","P0","Gujarati pricing"),
    ("/gu/almm-list-shu-che","ALMM List Shu Che — Gujarati","almm list gujarati",150,"Easy","I","P1","Gujarati technical"),
    ("/gu/net-metering-gujarat","Gujarat ma Net Metering Process","net metering gujarat gujarati",400,"Easy","I","P0","Gujarat-specific"),
    ("/gu/solar-warranty-gujarati","Solar Inverter Warranty — Gujarati","solar warranty gujarati",200,"Easy","I","P1","Gujarati trust"),
    ("/gu/whatsapp-solar-alerts","WhatsApp Solar Alert — Gujarati","whatsapp solar alert gujarati",80,"Easy","C","P1","USP Gujarati"),
    *[(f"/gu/solar-inverter-{c}", f"{c.replace('-',' ').title()} ma Solar Inverter — Gujarati", f"solar inverter {c.replace('-',' ')} gujarati", 400, "Easy", "L", "P0", f"Gujarat city Gujarati: {c.title()}") for c in ["ahmedabad","surat","rajkot","vadodara","gandhinagar","bhavnagar","jamnagar","junagadh","anand","bharuch","navsari","mehsana","valsad","nadiad","porbandar"]],
    ("/gu/best-solar-inverter-gujarat","Gujarat nu Best Solar Inverter 2026","best solar inverter gujarat gujarati",600,"Medium","C","P0","Gujarati BOFU"),
    ("/gu/atta-chakki-solar","Atta Chakki mate Solar Inverter","atta chakki solar gujarati",100,"Easy","C","P1","Tier-3 Gujarati"),
    ("/gu/pani-pump-solar","Pani Pump mate Solar Inverter — KUSUM","pani pump solar gujarati",300,"Easy","C","P0","Ag Gujarati"),
    ("/gu/dukaan-solar","Dukaan mate Solar Inverter","dukaan solar gujarati",200,"Easy","C","P0","SMB Gujarati"),
    ("/gu/factory-solar","Factory mate Solar System","factory solar gujarati",250,"Easy","C","P0","C&I Gujarati"),
    ("/gu/qbits-vs-luminous","Qbits vs Luminous — Gujarati","qbits vs luminous gujarati",80,"Easy","C","P0","Gujarati compare"),
    ("/gu/qbits-vs-microtek","Qbits vs Microtek — Gujarati","qbits vs microtek gujarati",60,"Easy","C","P0","Gujarati compare"),
    ("/gu/solar-inverter-chalu-nathi","Solar Inverter Chalu Nathi — Solution","solar inverter chalu nathi",200,"Easy","I","P1","Gujarati support"),
    ("/gu/inverter-garam-thay","Solar Inverter Garam Thay Che — Solution","inverter garam gujarati",150,"Easy","I","P1","Gujarati support"),
]
add_sheet("16_Gujarati_Content", ["URL","Meta Title","Target Keyword","Vol (mo)","KD","Intent","Priority","Notes"], guj, [50, 55, 32, 10, 10, 8, 9, 50])

# ========== 17 AI / WHATSAPP SPECIAL ==========
ai_wa = [
    ("/whatsapp-solar-support","WhatsApp Solar Monitoring & Alerts — Qbits USP","whatsapp solar alert",200,"Easy","C","P0","Uncontested SERP"),
    ("/ai-solar-inverter","AI-Powered Solar Inverter — How Qbits Works","ai solar inverter india",400,"Easy","C","P0","Qbits USP"),
    ("/ai-grid-tuning","AI Grid Tuning for Indian DISCOMs","ai grid tuning solar",100,"Easy","I","P0","Made-for-India tech"),
    ("/qbits-app","Qbits Mobile App — Monitor Your Solar Live","solar monitoring app india",1500,"Medium","C","P0","App download"),
    ("/whatsapp-bill-upload","Send Your Bill on WhatsApp, Get Instant Quote","solar quote whatsapp",300,"Easy","T","P0","Lead capture"),
    ("/predictive-fault-detection","Predictive Fault Detection — AI in Action","predictive maintenance solar",150,"Easy","I","P1","Tech narrative"),
    ("/qbits-vs-traditional-monitoring","Why WhatsApp Monitoring Beats Traditional App-Only","solar monitoring comparison",80,"Easy","I","P1","Comparison angle"),
    ("/whatsapp-installer-channel","WhatsApp Channel for Qbits Installers","solar installer whatsapp",100,"Easy","I","P1","B2B"),
    ("/multilingual-alerts","Multilingual Solar Alerts — 6 Indian Languages","multilingual solar alert",80,"Easy","C","P1","Inclusivity"),
    ("/founder-on-ai-strategy","Why We Built AI-First, Not AI-Bolted-On","ai solar story india",80,"Easy","I","P1","Founder POV"),
]
add_sheet("17_AI_WhatsApp_Special", ["URL","Meta Title","Target Keyword","Vol (mo)","KD","Intent","Priority","Notes"], ai_wa, [40, 55, 32, 10, 10, 8, 9, 50])

# ========== 18 MADE IN INDIA HUB ==========
mii = [
    ("/made-for-india","Made for India — Qbits Engineering Hub","made in india solar inverter",800,"Medium","I","P0","Hub home"),
    ("/made-for-india/bis-certification","BIS Certification Explained for Solar Inverters","bis certification solar inverter",1000,"Medium","I","P0","Trust pillar"),
    ("/made-for-india/almm-listing","Qbits on the ALMM List — What It Means","almm listing solar inverter",1500,"Medium","I","P0","Subsidy + trust"),
    ("/made-for-india/manufacturing","Inside Qbits' Indian Manufacturing Facility","qbits manufacturing india",100,"Easy","I","P1","Factory tour"),
    ("/made-for-india/data-sovereignty","Data Sovereignty — Where Your Solar Data Lives","data sovereignty solar",80,"Easy","I","P1","Privacy angle"),
    ("/made-for-india/german-igbt-india-assembly","German IGBTs, Assembled in India — Why It Matters","made in india igbt",60,"Easy","I","P1","Component story"),
    ("/made-for-india/atmanirbhar-bharat","Qbits & Atmanirbhar Bharat","atmanirbhar bharat solar",200,"Easy","I","P0","Patriotic narrative"),
    ("/made-for-india/iso-9001","Qbits ISO 9001 Quality Certification","iso 9001 solar inverter",150,"Easy","I","P1","Trust"),
    ("/made-for-india/iso-14001","Qbits ISO 14001 Environmental Certification","iso 14001 solar inverter",100,"Easy","I","P2","ESG"),
    ("/made-for-india/govt-tender-bidding","Qbits in Government Solar Tenders","govt solar tender india",300,"Easy","C","P1","B2G"),
    ("/made-for-india/psu-installations","Qbits at PSU & Govt Installations","psu solar inverter",80,"Easy","C","P2","Reference list"),
    ("/made-for-india/import-vs-india","Why Import Solar Brands Struggle in India","imported solar inverter india",200,"Easy","I","P1","Competitive frame"),
    ("/made-for-india/quality-policy","Qbits Quality Policy & Testing Protocols","solar inverter quality testing",150,"Easy","I","P2","Trust depth"),
    ("/made-for-india/rohs-compliance","RoHS Compliance for Solar Inverters","rohs solar inverter",100,"Easy","I","P2","Compliance"),
    ("/made-for-india/csr-sustainability","Qbits CSR & Sustainability Commitments","solar csr india",80,"Easy","I","P2","ESG"),
]
add_sheet("18_Made_In_India", ["URL","Meta Title","Target Keyword","Vol (mo)","KD","Intent","Priority","Notes"], mii, [42, 55, 32, 10, 10, 8, 9, 50])

# ========== KEYWORD TABS ==========

# KW_BOFU
bofu = [
    ("qbits 5kw inverter price","Branded SKU","Bottom","80","Easy","Product"),
    ("qbits hybrid inverter","Branded","Bottom","100","Easy","Product"),
    ("solar inverter dealer near me","Local","Bottom","2500","Hard","Dealer-Locator"),
    ("solar installer near me","Local","Bottom","5000","Hard","Tool"),
    ("buy solar inverter online india","Transactional","Bottom","1500","Hard","Product"),
    ("solar inverter price with installation","Pricing","Bottom","700","Medium","Pricing/Calc"),
    ("solar inverter on emi","Pricing","Bottom","1500","Medium","EMI Calc"),
    ("solar inverter dealer price","Pricing","Bottom","600","Medium","Dealer Form"),
    ("5kw solar inverter price","Pricing","Bottom","6000","Hard","Product 5kW"),
    ("3kw solar inverter price","Pricing","Bottom","4000","Hard","Product 3kW"),
    ("10kw solar inverter price","Pricing","Bottom","2000","Medium","Product 10kW"),
    ("5kw hybrid inverter price india","Pricing","Bottom","2500","Hard","Product QH5"),
    ("solar inverter with subsidy","Subsidy intent","Bottom","800","Medium","PM Surya Ghar"),
    ("pm surya ghar subsidy amount","Subsidy","Bottom","12000","Hard","Subsidy Hub"),
    ("pm surya ghar apply online","Subsidy","Bottom","45000","Very Hard","PM Surya Ghar"),
    ("solar inverter for home price","Pricing","Bottom","4000","Hard","Home use-case"),
    ("hybrid inverter price india","Pricing","Bottom","3500","Hard","Hybrid hub"),
    ("on grid inverter price india","Pricing","Bottom","3000","Hard","On-grid hub"),
    ("best solar inverter for home","Commercial","Bottom","6000","Very Hard","Best-list"),
    ("best 5kw solar inverter","Commercial","Bottom","4000","Hard","Best 5kW"),
    ("best hybrid inverter india","Commercial","Bottom","2500","Hard","Best hybrid"),
    ("qbits vs growatt","Comparison","Bottom","150","Easy","Compare page"),
    ("qbits vs luminous","Comparison","Bottom","200","Easy","Compare page"),
    ("qbits vs sungrow","Comparison","Bottom","100","Easy","Compare page"),
    ("qbits vs deye","Comparison","Bottom","100","Easy","Compare page"),
    ("solar inverter installation cost","Pricing","Bottom","2500","Hard","Install pricing"),
    ("solar inverter for shop","Use-case","Bottom","1500","Medium","Shop vertical"),
    ("solar inverter for factory","Use-case","Bottom","1000","Medium","Factory vertical"),
    ("solar inverter for water pump","Use-case","Bottom","2500","Hard","Pump (KUSUM)"),
    ("solar inverter complaint number","Support","Bottom","800","Easy","Service page"),
]
add_sheet("KW_BOFU", ["Keyword","Type","Funnel","Vol (mo)","KD","Page to Target"], bofu, [40, 18, 12, 12, 12, 28])

# KW_Commercial (mid-funnel)
commercial = [
    ("best solar inverter india",4500,"Very Hard","C","Best-list pillar"),
    ("best solar inverter brands india",2500,"Hard","C","Best-list"),
    ("top 10 solar inverter india",3500,"Hard","C","Best-list"),
    ("hybrid solar inverter india",12000,"Hard","C","Hybrid hub"),
    ("on grid solar inverter india",6000,"Hard","C","On-grid hub"),
    ("off grid solar inverter india",5000,"Hard","C","Off-grid hub"),
    ("3 phase solar inverter india",2500,"Medium","C","3P hub"),
    ("smart solar inverter india",800,"Medium","C","AI page"),
    ("wifi solar inverter india",1500,"Medium","C","Wi-Fi review"),
    ("ai solar inverter india",400,"Easy","C","Qbits AI page"),
    ("mppt solar inverter india",2500,"Medium","C","Glossary+Product"),
    ("hybrid inverter for home",4000,"Hard","C","Home + hybrid"),
    ("solar inverter for 3 bhk",1500,"Medium","C","3-BHK best list"),
    ("solar inverter for villa",600,"Medium","C","Villa best list"),
    ("solar inverter for ev charging",500,"Medium","C","EV use-case"),
    ("indian solar inverter brands",1500,"Medium","C","Best Indian brands"),
    ("which solar inverter is best",2000,"Hard","C","Buying guide"),
    ("most reliable solar inverter",500,"Medium","C","Best reliable"),
    ("solar inverter with longest warranty",400,"Easy","C","Best-warranty page"),
    ("solar inverter app monitoring",500,"Medium","C","App page"),
    ("almm listed inverters",2500,"Hard","C","ALMM hub"),
    ("solar inverter ip65",400,"Easy","C","Glossary IP rating"),
    ("dual mppt solar inverter",600,"Medium","C","Glossary MPPT"),
    ("transformerless solar inverter",800,"Medium","C","Glossary + Best"),
    ("hybrid vs on grid inverter",2200,"Hard","C","Comparison page"),
    ("string vs micro inverter",1800,"Hard","C","Comparison page"),
    ("solar inverter for ac",1000,"Medium","C","Load use-case blog"),
    ("solar system for 3 bhk",2000,"Hard","C","Use-case page"),
    ("solar inverter wholesale price",600,"Medium","C","Dealer form"),
    ("solar inverter buy online india",1500,"Hard","T","Product hub"),
]
add_sheet("KW_Commercial", ["Keyword","Vol (mo)","KD","Intent","Page to Target"], commercial, [40, 12, 12, 10, 32])

# KW_Informational
info = [
    ("how does solar inverter work",7000,"Hard","Foundational blog"),
    ("what is mppt",2500,"Medium","Glossary"),
    ("how to choose solar inverter",1500,"Medium","Buying-guide blog"),
    ("solar inverter working principle",2000,"Medium","Foundational blog"),
    ("difference between on grid and off grid",3000,"Hard","Comparison page"),
    ("what is hybrid inverter",2000,"Medium","Glossary"),
    ("solar inverter sizing",1500,"Medium","Sizing blog/calc"),
    ("solar inverter efficiency",800,"Medium","Glossary"),
    ("solar inverter components",600,"Medium","Foundational"),
    ("solar inverter circuit diagram",1500,"Medium","Diagram blog"),
    ("solar inverter wiring diagram",1500,"Medium","Wiring blog"),
    ("solar inverter life span",800,"Medium","Lifecycle blog"),
    ("solar inverter maintenance",1500,"Medium","Maintenance blog"),
    ("solar inverter installation guide",800,"Medium","Installation blog"),
    ("what is thd in inverter",500,"Medium","Glossary"),
    ("what is ip65 rating",800,"Medium","Glossary"),
    ("solar inverter buying guide",1500,"Medium","Pillar blog"),
    ("how to read inverter datasheet",400,"Easy","Datasheet blog"),
    ("solar panel degradation",600,"Medium","Degradation blog"),
    ("what is anti islanding",300,"Easy","Glossary"),
    ("solar inverter heat issue",400,"Easy","Troubleshoot blog"),
    ("how to size solar inverter for home",800,"Medium","Sizing blog"),
    ("mppt vs pwm",1500,"Medium","Comparison"),
    ("solar inverter vs ups",1200,"Medium","Comparison"),
    ("what is net metering",2500,"Medium","Net-metering blog"),
    ("inverter battery connection",1500,"Medium","Wiring blog"),
    ("solar inverter app explained",400,"Easy","App blog"),
    ("solar yield india",300,"Easy","Yield blog"),
    ("payback period solar",800,"Medium","Calc + blog"),
    ("lcoe solar india",200,"Easy","Tech blog"),
]
add_sheet("KW_Informational", ["Keyword","Vol (mo)","KD","Page to Target"], info, [40, 12, 12, 32])

# KW_Hindi
hindi_kw = [
    ("pm surya ghar yojana hindi",30000,"Hard","Hindi scheme hub"),
    ("solar subsidy kaise milegi",4000,"Medium","Hindi subsidy guide"),
    ("solar inverter kya hai",4000,"Medium","Hindi info"),
    ("ghar me solar panel kaise lagaye",2500,"Medium","Hindi how-to"),
    ("solar panel ki kimat",2000,"Medium","Hindi pricing"),
    ("bijli bill kaise kam kare",5000,"Medium","Hindi top funnel"),
    ("mppt kya hota hai",2500,"Easy","Hindi glossary"),
    ("on grid off grid antar",1500,"Easy","Hindi comparison"),
    ("ghar ke liye solar inverter",3000,"Medium","Hindi BOFU"),
    ("5kw solar inverter price hindi",2000,"Medium","Hindi product"),
    ("3kw solar inverter price hindi",1500,"Medium","Hindi product"),
    ("hybrid solar inverter hindi",1200,"Easy","Hindi info"),
    ("solar inverter installation cost hindi",800,"Easy","Hindi install"),
    ("solar inverter warranty hindi",500,"Easy","Hindi trust"),
    ("solar inverter band ho gaya",1500,"Easy","Hindi support"),
    ("solar inverter ups antar",1000,"Easy","Hindi info"),
    ("anti islanding hindi",300,"Easy","Hindi glossary"),
    ("kusum yojana hindi",8000,"Medium","Hindi farmer"),
    ("subsidy ke baad solar inverter price",500,"Easy","Hindi subsidy pricing"),
    ("atta chakki ke liye solar",500,"Easy","Hindi vertical"),
    ("petrol pump ke liye solar",300,"Easy","Hindi vertical"),
    ("solar inverter chalu nahi ho raha",1500,"Easy","Hindi support"),
    ("solar inverter battery kab badle",400,"Easy","Hindi maintenance"),
    ("inverter wifi connect nahi ho raha",800,"Easy","Hindi support"),
    ("uttar pradesh solar subsidy hindi",1500,"Medium","UP hindi"),
    ("bihar solar subsidy hindi",800,"Medium","Bihar hindi"),
    ("mp solar subsidy hindi",600,"Medium","MP hindi"),
    ("rajasthan solar subsidy hindi",800,"Medium","Rajasthan hindi"),
    ("ip65 rating ka matlab hindi",200,"Easy","Hindi glossary"),
    ("solar inverter ki life hindi",400,"Easy","Hindi lifespan"),
    ("ghar ke liye solar inverter kaise size karein",800,"Medium","Hindi sizing"),
    ("solar panel ke saath inverter",600,"Easy","Hindi combo"),
    ("solar panel ka subsidy form",1500,"Medium","Hindi subsidy form"),
    ("solar installation kahan se karaye",2500,"Medium","Hindi dealer"),
    ("bijli bill zero kaise kare",2000,"Medium","Hindi top funnel"),
    ("solar inverter ka rate",1500,"Medium","Hindi pricing"),
    ("solar panel lagane ka kharcha",1800,"Medium","Hindi cost"),
    ("solar inverter customer care",500,"Easy","Hindi support"),
    ("solar ka subsidy kab milta hai",800,"Easy","Hindi timeline"),
    ("solar ka loan",800,"Easy","Hindi EMI"),
]
add_sheet("KW_Hindi", ["Keyword","Vol (mo)","KD","Page to Target"], hindi_kw, [42, 12, 12, 32])

# KW_Subsidy
sub_kw = [
    ("pm surya ghar yojana",200000,"Very Hard","PM Surya Ghar hub"),
    ("pm surya ghar registration",60000,"Very Hard","PM Surya Ghar hub"),
    ("pm surya ghar apply online",45000,"Very Hard","Apply guide"),
    ("pm surya ghar status",30000,"Hard","Status check"),
    ("solar subsidy 2026",30000,"Very Hard","Subsidy hub"),
    ("rooftop solar subsidy",25000,"Very Hard","MNRE rooftop"),
    ("free solar panel scheme",18000,"Very Hard","PM Surya Ghar"),
    ("pm surya ghar eligibility",20000,"Hard","Eligibility blog"),
    ("pm surya ghar subsidy amount",15000,"Hard","Calculator"),
    ("pm surya ghar documents",8000,"Hard","Docs page"),
    ("pm surya ghar 2026",10000,"Hard","Hub"),
    ("kusum yojana",30000,"Very Hard","KUSUM hub"),
    ("kusum scheme apply",8000,"Hard","KUSUM apply"),
    ("mnre subsidy solar",8000,"Hard","MNRE hub"),
    ("mnre rooftop scheme",3500,"Hard","Scheme hub"),
    ("almm list 2026",2500,"Hard","ALMM blog"),
    ("almm phase 3",1500,"Medium","ALMM update"),
    ("net metering india",4000,"Hard","Net metering blog"),
    ("gross metering india",1500,"Medium","Comparison blog"),
    ("net metering process",4000,"Hard","Process blog"),
    ("solar subsidy gujarat",3000,"Medium","Gujarat hub"),
    ("solar subsidy maharashtra",3500,"Medium","MH hub"),
    ("solar subsidy tamil nadu",2500,"Medium","TN hub"),
    ("solar subsidy karnataka",2000,"Medium","KA hub"),
    ("solar subsidy uttar pradesh",2500,"Medium","UP hub"),
    ("solar subsidy rajasthan",2000,"Medium","RJ hub"),
    ("solar subsidy kerala",1500,"Medium","KL hub"),
    ("solar subsidy delhi",4000,"Hard","DL hub"),
    ("solar subsidy west bengal",1500,"Medium","WB hub"),
    ("solar subsidy bihar",1500,"Medium","BR hub"),
    ("solar subsidy mp",1500,"Medium","MP hub"),
    ("surya gujarat scheme",2500,"Medium","Surya Gujarat blog"),
    ("kusum yojana online application",3000,"Hard","KUSUM apply blog"),
    ("ireda loan solar",1200,"Medium","Finance blog"),
    ("sbi solar loan",1500,"Medium","Finance blog"),
    ("solar panel subsidy form download",2000,"Hard","Form download"),
    ("discom empanelled vendor list",1500,"Medium","Empanelled blog"),
    ("solar net meter approval time",800,"Medium","Timeline blog"),
    ("gst on solar 2026",1500,"Medium","GST blog"),
    ("bcd on solar inverter",400,"Easy","Customs blog"),
]
add_sheet("KW_Subsidy", ["Keyword","Vol (mo)","KD","Page to Target"], sub_kw, [42, 12, 12, 32])

# KW_Comparison
cmp_kw = [
    ("sungrow vs growatt",1200,"Medium","Compare"),
    ("growatt vs solis",700,"Medium","Compare"),
    ("sungrow vs solis",600,"Medium","Compare"),
    ("luminous vs microtek",1500,"Medium","Compare"),
    ("luminous vs sukam",800,"Medium","Compare"),
    ("luminous vs havells",700,"Medium","Compare"),
    ("microtek vs sukam",500,"Medium","Compare"),
    ("havells vs polycab",400,"Medium","Compare"),
    ("deye vs growatt",600,"Medium","Compare"),
    ("waaree vs adani",700,"Medium","Compare"),
    ("goodwe vs growatt",500,"Medium","Compare"),
    ("utl vs luminous",600,"Medium","Compare"),
    ("fronius vs sma",400,"Medium","Compare"),
    ("huawei vs sungrow",400,"Medium","Compare"),
    ("enphase vs solaredge",500,"Medium","Compare"),
    ("indian vs chinese solar inverter",400,"Easy","Compare"),
    ("hybrid vs on grid inverter",2200,"Hard","Compare"),
    ("on grid vs off grid",3500,"Hard","Compare"),
    ("string vs micro inverter",1800,"Hard","Compare"),
    ("mppt vs pwm",1500,"Medium","Compare"),
    ("single phase vs 3 phase inverter",800,"Medium","Compare"),
    ("transformer vs transformerless",600,"Medium","Compare"),
    ("solar inverter vs ups",1200,"Medium","Compare"),
    ("5kw vs 10kw inverter",400,"Easy","Compare"),
    ("qbits vs growatt",150,"Easy","Compare"),
    ("qbits vs luminous",200,"Easy","Compare"),
    ("qbits vs sungrow",100,"Easy","Compare"),
    ("qbits vs deye",100,"Easy","Compare"),
    ("qbits vs microtek",100,"Easy","Compare"),
    ("qbits vs utl",80,"Easy","Compare"),
]
add_sheet("KW_Comparison", ["Keyword","Vol (mo)","KD","Page to Target"], cmp_kw, [40, 12, 12, 28])

# KW_Geo_Local
geo_kw = []
for c in tier1:
    geo_kw.append([f"solar inverter {c}", 2500, "Medium", "Local", f"City page {c}"])
    geo_kw.append([f"best solar inverter {c}", 1500, "Medium", "Local", f"Best-list {c}"])
    geo_kw.append([f"solar installer {c}", 1500, "Medium", "Local", f"Dealer locator {c}"])
for c in tier2[:20]:
    geo_kw.append([f"solar inverter {c.replace('-',' ')}", 900, "Medium", "Local", f"City page {c}"])
    geo_kw.append([f"solar dealer {c.replace('-',' ')}", 400, "Easy", "Local", f"Dealer locator"])
for s in priority_states:
    geo_kw.append([f"solar inverter {s.replace('-',' ')}", 1500, "Medium", "Local", f"State page {s}"])
    geo_kw.append([f"best solar inverter {s.replace('-',' ')}", 900, "Medium", "Local", f"Best-list {s}"])
add_sheet("KW_Geo_Local", ["Keyword","Vol (mo)","KD","Intent","Page to Target"], geo_kw, [40, 12, 12, 10, 32])

# KW_LongTail_Patterns
patterns = [
    ("{wattage} solar inverter price {city}","30 wattages × 50 cities = 1,500 variants","Use programmatic SEO"),
    ("{wattage} {type} inverter {price-qualifier}","30 × 4 × 8 = 960 variants","Pricing pages"),
    ("{brand} vs {brand} {wattage}","20 brands × 19 × 10 wattages = ~3,800","Comparison pages"),
    ("solar inverter for {use-case} in {city}","30 use × 50 cities = 1,500","City use-case pages"),
    ("solar subsidy {state} {year}","28 states × 3 years = 84","Subsidy hub variants"),
    ("solar inverter {error-code} solution","20 codes × 4 SKUs = 80","Error library"),
    ("{brand} {model} review india","80 brand/models","Review pages"),
    ("how to {action} solar inverter","20 actions","How-to blogs"),
    ("what is {term} in solar inverter","150 glossary terms","Glossary"),
    ("solar inverter {language-suffix}","Hindi/Gujarati/Tamil/Telugu variants","Regional pages"),
    ("pm surya ghar {state}","28 states","State subsidy"),
    ("kusum scheme {state}","28 states","KUSUM state pages"),
    ("net metering {discom}","40+ DISCOMs","Installer paperwork"),
    ("solar inverter for {industry-vertical} in {state}","30 verticals × 10 states = 300","C&I vertical pages"),
    ("best solar inverter under {price}","6 price points","Best-under pages"),
]
add_sheet("KW_LongTail_Patterns", ["Pattern","Scale (rough)","Generation Method"], patterns, [55, 50, 38])

# Save
out = "/home/ritik/qbits-energy-website/Marketing/Qbits_Website_Master_Workbook.xlsx"
wb.save(out)
print(f"SAVED: {out}")
print(f"Sheets: {len(wb.sheetnames)}")
for s in wb.sheetnames: print(f"  - {s}")
